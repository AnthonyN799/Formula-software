import streamlit as st
import pandas as pd
from datetime import datetime
from supabase import create_client, Client
from PIL import Image
import re
import time
from fpdf import FPDF

# --- 1. PAGE CONFIGURATION ---
try:
    logo_img = Image.open("logo.jpg")
    st.set_page_config(page_title="Therapeutic Oils | Lab Portal", page_icon=logo_img, layout="wide", initial_sidebar_state="expanded")
except FileNotFoundError:
    st.set_page_config(page_title="Therapeutic Oils | Lab Portal", layout="wide", initial_sidebar_state="expanded")

# --- 2. CUSTOM CSS ---
def inject_custom_css():
    st.markdown("""
        <style>
        #MainMenu {visibility: hidden;} footer {visibility: hidden;} header {background-color: transparent !important;}
        .stApp { font-family: 'Inter', -apple-system, sans-serif; }
        [data-testid="stMetricValue"] { font-size: 2.2rem; font-weight: 300; letter-spacing: -0.02em; }
        [data-testid="stMetricLabel"] { font-size: 0.85rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }
        [data-testid="metric-container"] { padding: 1.5rem; border-radius: 8px; }
        .stButton>button { border-radius: 4px; font-weight: 500; transition: all 0.2s ease; }
        h1, h2, h3 { font-weight: 400; letter-spacing: -0.01em; }
        /* Custom minimal spinner */
        .stSpinner > div { display: flex; align-items: center; justify-content: center; }
        .stSpinner > div > div { width: 20px !important; height: 20px !important; border-width: 2px !important; }
        .stSpinner > div > span { font-size: 0.85rem !important; font-weight: 500 !important; letter-spacing: 0.02em !important; margin-left: 8px !important; }
        /* Hide default "Running..." top-right badge */
        [data-testid="stStatusWidget"] { display: none !important; }
        /* Ensure markdown text inherits theme color */
        .stMarkdown p, .stMarkdown span, .stMarkdown li { color: inherit !important; }
        </style>
    """, unsafe_allow_html=True)

# --- Connect to the Database ---
@st.cache_resource
def init_connection():
    return create_client(st.secrets["supabase"]["url"], st.secrets["supabase"]["key"])
supabase = init_connection()

@st.cache_data(ttl=60, show_spinner=False)
def _fetch_cached(table_name, sort_column=None):
    for attempt in range(3): 
        try:
            resp = supabase.table(table_name).select("*").execute()
            df = pd.DataFrame(resp.data) if resp.data else pd.DataFrame()
            if not df.empty and sort_column and sort_column in df.columns:
                df = df.sort_values(sort_column)
            return df
        except Exception: time.sleep(0.5) 
    return pd.DataFrame()

def fetch_vault_data(table_name, sort_column=None):
    df = _fetch_cached(table_name, sort_column)
    if df is None:
        st.error(f"⚠️ Network timeout accessing {table_name}. Please refresh.")
        st.stop()
    return df.copy()

def clear_cache():
    _fetch_cached.clear()

def load_tables(*names):
    table_map = {
        'inventory': ('inventory', 'rm_code'),
        'packaging': ('packaging', 'pm_code'),
        'finished_goods': ('finished_products', 'fp_code'),
        'formulas': ('formulas', None),
        'cogs_records': ('cogs_records', 'product_name'),
        'sales_records': ('sales_records', 'sale_date'),
        'consignment': ('consignment_records', 'created_at'),
        'clients': ('clients', 'client_name'),
        'portfolios': ('portfolios', 'portfolio_name'),
    }
    with st.spinner("Loading..."):
        result = {}
        for name in names:
            tbl, sort = table_map[name]
            result[name] = fetch_vault_data(tbl, sort)
    return result

# --- PDF Engines ---
def generate_order_pdf(order_ref, items_df, client_name, date_str):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 18)
    pdf.cell(0, 10, "THERAPEUTIC OILS", ln=True, align="C")
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, "Official Order Summary", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(100, 8, f"Billed To: {client_name}")
    pdf.cell(0, 8, f"Date: {date_str}", ln=True, align="R")
    pdf.cell(0, 8, f"Order Ref: ORD-{(int(order_ref) + 200):06d}" if str(order_ref).isdigit() else f"Order Ref: {order_ref}", ln=True, align="R")
    pdf.ln(5)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(90, 8, "Product Description", border=1)
    pdf.cell(25, 8, "Qty", border=1, align="C")
    pdf.cell(35, 8, "Unit Price", border=1, align="R")
    pdf.cell(40, 8, "Line Total", border=1, align="R")
    pdf.ln()
    pdf.set_font("Arial", "", 10)
    grand_total = 0.0
    for _, row in items_df.iterrows():
        total = float(row['gross_revenue'])
        grand_total += total
        pdf.cell(90, 8, str(row['order_description'])[:45], border=1)
        pdf.cell(25, 8, str(row['quantity']), border=1, align="C")
        pdf.cell(35, 8, f"${float(row['unit_price']):,.2f}", border=1, align="R")
        pdf.cell(40, 8, f"${total:,.2f}", border=1, align="R")
        pdf.ln()
    pdf.set_font("Arial", "B", 10)
    pdf.cell(150, 8, "Grand Total", border=1, align="R")
    pdf.cell(40, 8, f"${grand_total:,.2f}", border=1, align="R")
    return pdf.output(dest="S").encode("latin-1")

def generate_consignment_pdf(order_ref, items_df, partner_name, date_str):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 18)
    pdf.cell(0, 10, "THERAPEUTIC OILS", ln=True, align="C")
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, "Official Consignment Agreement", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(100, 8, f"Consignee (Partner): {partner_name}")
    pdf.cell(0, 8, f"Date Issued: {date_str}", ln=True, align="R")
    pdf.cell(0, 8, f"Reference #: {order_ref}", ln=True, align="R")
    pdf.ln(5)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(75, 8, "Product Description", border=1)
    pdf.cell(20, 8, "Qty Sent", border=1, align="C")
    pdf.cell(30, 8, "Retail Price", border=1, align="R")
    pdf.cell(30, 8, "Owed to Maker", border=1, align="R")
    pdf.cell(35, 8, "Max Potential", border=1, align="R")
    pdf.ln()
    pdf.set_font("Arial", "", 9)
    grand_total_owed = 0.0
    for _, row in items_df.iterrows():
        wholesale = float(row['wholesale_price'])
        total_owed = float(row['qty_consigned'] * wholesale)
        grand_total_owed += total_owed
        pdf.cell(75, 8, str(row['product_name'])[:35], border=1)
        pdf.cell(20, 8, str(row['qty_consigned']), border=1, align="C")
        pdf.cell(30, 8, f"${float(row['retail_price']):,.2f}", border=1, align="R")
        pdf.cell(30, 8, f"${wholesale:,.2f}", border=1, align="R")
        pdf.cell(35, 8, f"${total_owed:,.2f}", border=1, align="R")
        pdf.ln()
    pdf.ln(5)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(155, 8, "Total Capital Owed Upon 100% Sell-Through:", border=1, align="R")
    pdf.cell(35, 8, f"${grand_total_owed:,.2f}", border=1, align="R")
    pdf.ln(15)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Terms of Consignment:", ln=True)
    pdf.set_font("Arial", "", 9)
    terms = (
        "1. Title and ownership of all goods listed above remain strictly with Therapeutic Oils until sold to an end consumer.\n"
        "2. The Consignee agrees to display and store the goods appropriately.\n"
        "3. The 'Owed to Maker' amount must be paid to Therapeutic Oils for every unit sold during the reporting period.\n"
        "4. Unsold goods may be recalled by Therapeutic Oils or returned by the Consignee at any time, provided they are in sellable condition."
    )
    pdf.multi_cell(0, 5, terms)
    return pdf.output(dest="S").encode("latin-1")

def generate_partner_inventory_pdf(partner_name, items_df, date_str):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 18)
    pdf.cell(0, 10, "THERAPEUTIC OILS", ln=True, align="C")
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, "Partner Consignment Statement", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(100, 8, f"Partner: {partner_name}")
    pdf.cell(0, 8, f"Date: {date_str}", ln=True, align="R")
    pdf.ln(5)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(45, 8, "Product", border=1)
    pdf.cell(25, 8, "Ref #", border=1)
    pdf.cell(18, 8, "Consigned", border=1, align="C")
    pdf.cell(15, 8, "Sold", border=1, align="C")
    pdf.cell(20, 8, "Remaining", border=1, align="C")
    pdf.cell(22, 8, "Unit Price", border=1, align="R")
    pdf.cell(23, 8, "Total Price", border=1, align="R")
    pdf.cell(22, 8, "Retail/Unit", border=1, align="R")
    pdf.ln()
    pdf.set_font("Arial", "", 9)
    total_remaining = 0
    total_owed = 0.0
    for _, row in items_df.iterrows():
        remaining = int(row['qty_consigned']) - int(row['qty_sold'])
        owed = float(remaining) * float(row['wholesale_price'])
        total_remaining += remaining
        total_owed += owed
        pdf.cell(45, 8, str(row['product_name'])[:23], border=1)
        pdf.cell(25, 8, str(row['order_ref_number'])[:12], border=1)
        pdf.cell(18, 8, str(int(row['qty_consigned'])), border=1, align="C")
        pdf.cell(15, 8, str(int(row['qty_sold'])), border=1, align="C")
        pdf.cell(20, 8, str(remaining), border=1, align="C")
        pdf.cell(22, 8, f"${float(row['wholesale_price']):,.2f}", border=1, align="R")
        pdf.cell(23, 8, f"${owed:,.2f}", border=1, align="R")
        pdf.cell(22, 8, f"${float(row['retail_price']):,.2f}", border=1, align="R")
        pdf.ln()
    pdf.ln(5)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(167, 8, "Total Capital Owed Upon 100% Sell-Through:", border=1, align="R")
    pdf.cell(23, 8, f"${total_owed:,.2f}", border=1, align="R")
    pdf.ln(15)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Terms of Consignment:", ln=True)
    pdf.set_font("Arial", "", 9)
    terms = (
        "1. Title and ownership of all goods listed above remain strictly with Therapeutic Oils until sold to an end consumer.\n"
        "2. The Consignee agrees to display and store the goods appropriately.\n"
        "3. The 'Total Price' amount must be paid to Therapeutic Oils for every unit sold during the reporting period.\n"
        "4. Unsold goods may be recalled by Therapeutic Oils or returned by the Consignee at any time, provided they are in sellable condition."
    )
    pdf.multi_cell(0, 5, terms)
    return pdf.output(dest="S").encode("latin-1")

def generate_balance_sheet_pdf(date_str, cash, ar, inv_rm, inv_pm, inv_fg, fixed_assets, ap, debt, total_assets, total_liab, equity):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 18)
    pdf.cell(0, 10, "THERAPEUTIC OILS", ln=True, align="C")
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, "Balance Sheet", ln=True, align="C")
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 6, f"As of {date_str}", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "ASSETS", ln=True, border="B")
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Current Assets", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(140, 6, "Cash & Equivalents:")
    pdf.cell(0, 6, f"${cash:,.2f}", ln=True, align="R")
    pdf.cell(140, 6, "Accounts Receivable:")
    pdf.cell(0, 6, f"${ar:,.2f}", ln=True, align="R")
    pdf.cell(140, 6, "Inventory (Raw Materials):")
    pdf.cell(0, 6, f"${inv_rm:,.2f}", ln=True, align="R")
    pdf.cell(140, 6, "Inventory (Packaging):")
    pdf.cell(0, 6, f"${inv_pm:,.2f}", ln=True, align="R")
    pdf.cell(140, 6, "Inventory (Finished Goods):")
    pdf.cell(0, 6, f"${inv_fg:,.2f}", ln=True, align="R")
    pdf.ln(2)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Fixed Assets", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(140, 6, "Property, Plant & Equipment:")
    pdf.cell(0, 6, f"${fixed_assets:,.2f}", ln=True, align="R")
    pdf.ln(2)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(140, 8, "TOTAL ASSETS:")
    pdf.cell(0, 8, f"${total_assets:,.2f}", ln=True, align="R")
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "LIABILITIES & EQUITY", ln=True, border="B")
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Liabilities", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(140, 6, "Accounts Payable (Unpaid Bills):")
    pdf.cell(0, 6, f"${ap:,.2f}", ln=True, align="R")
    pdf.cell(140, 6, "Short/Long Term Debt:")
    pdf.cell(0, 6, f"${debt:,.2f}", ln=True, align="R")
    pdf.ln(2)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(140, 6, "Total Liabilities:")
    pdf.cell(0, 6, f"${total_liab:,.2f}", ln=True, align="R")
    pdf.ln(4)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Owner's Equity", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.cell(140, 6, "Total Equity (Assets - Liabilities):")
    pdf.cell(0, 6, f"${equity:,.2f}", ln=True, align="R")
    pdf.ln(2)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(140, 8, "TOTAL LIABILITIES & EQUITY:")
    pdf.cell(0, 8, f"${(total_liab + equity):,.2f}", ln=True, align="R")
    return pdf.output(dest="S").encode("latin-1")

def generate_batch_labels_pdf(product_name, batch_number, lot_number, date_str, copies=6):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "THERAPEUTIC OILS - GMP BATCH LABELS", ln=True, align="C")
    pdf.ln(5)
    label_w = 90
    label_h = 45
    margin_x = 10
    margin_y = 30
    spacing_x = 10
    spacing_y = 10
    for i in range(copies):
        row = i // 2
        col = i % 2
        x = margin_x + col * (label_w + spacing_x)
        y = margin_y + row * (label_h + spacing_y)
        pdf.rect(x, y, label_w, label_h)
        pdf.set_xy(x, y + 5)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(label_w, 6, "THERAPEUTIC OILS", ln=True, align="C")
        pdf.set_xy(x, y + 13)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(label_w, 5, product_name[:35], ln=True, align="C")
        pdf.set_xy(x + 5, y + 22)
        pdf.set_font("Arial", "", 9)
        pdf.cell(label_w - 10, 5, f"BATCH: {batch_number}", ln=True)
        pdf.set_x(x + 5)
        pdf.cell(label_w - 10, 5, f"LOT: {lot_number}", ln=True)
        pdf.set_x(x + 5)
        pdf.cell(label_w - 10, 5, f"MFG DATE: {date_str}", ln=True)
        pdf.set_xy(x, y + label_h - 7)
        pdf.set_font("Arial", "I", 7)
        pdf.cell(label_w, 4, "Store in a cool, dark environment. Follow standard SOP.", ln=True, align="C")
    return pdf.output(dest="S").encode("latin-1")

# --- Authentication Logic ---
def check_password():
    if "authenticated" not in st.session_state: st.session_state["authenticated"] = False
    if st.session_state["authenticated"]: return True
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.write("<br><br><br>", unsafe_allow_html=True)
        try: st.image("logo.jpg", use_container_width=True)
        except: st.markdown("<h1 style='text-align: center; font-weight: 300;'>Therapeutic Oils</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; opacity: 0.6;'>Secure Portal</p>", unsafe_allow_html=True)
        username = st.text_input("Username", placeholder="Enter your name...")
        password = st.text_input("Passcode", type="password", placeholder="Enter passcode...")
        if st.button("Authenticate", use_container_width=True, type="primary"):
            users = {"anthony": {"pass": "lab2026", "role": "admin"}, "fadia": {"pass": "FadiaBoustany123", "role": "Analyst"}}
            matched = users.get(username.strip().lower())
            if matched and password == matched["pass"]:
                st.session_state["authenticated"] = True
                st.session_state["user_role"] = matched["role"]
                st.session_state["user_name"] = username
                st.rerun()
            else: st.error("Incorrect username or passcode.")
    return False

# --- Main App Execution ---
if check_password():
    inject_custom_css()

    # --- AUTO-SYNC COGS based on current RM avg prices ---
    # Smart versioning:
    #   - Drift > 5%: archive old, create new version (preserves audit trail for big changes)
    #   - Drift > $0.01 but < 5%: silent in-place update (no version spam)
    # Historical records (FP unit_cogs, sales_records) are NEVER touched.
    def auto_sync_cogs():
        try:
            cogs_resp = supabase.table('cogs_records').select("*").execute()
            cogs_records = pd.DataFrame(cogs_resp.data) if cogs_resp.data else pd.DataFrame()
            if cogs_records.empty: return
            if 'is_active' in cogs_records.columns:
                active = cogs_records[cogs_records['is_active'] != False]
            else:
                active = cogs_records
            if active.empty: return
            inv_resp = supabase.table('inventory').select("*").execute()
            inv_df = pd.DataFrame(inv_resp.data) if inv_resp.data else pd.DataFrame()
            form_resp = supabase.table('formulas').select("*").execute()
            form_df = pd.DataFrame(form_resp.data) if form_resp.data else pd.DataFrame()
            if inv_df.empty or form_df.empty: return

            def _rm_avg(mat_row):
                lots = mat_row.get('lots', [])
                if isinstance(lots, float) or (isinstance(lots, str) and lots in ["", "nan", "[]"]):
                    lots = []
                default_p = float(mat_row['price_per_kg'])
                if not lots: return default_p
                total_val = sum(float(l.get('Qty (Kg)', 0)) * float(l.get('Price/Kg', default_p)) for l in lots)
                total_q = sum(float(l.get('Qty (Kg)', 0)) for l in lots)
                return (total_val / total_q) if total_q > 0 else default_p

            updates = 0
            for _, prof in active.iterrows():
                fname = prof['formula_name']
                if fname == "None" or fname not in form_df['formula_name'].values: continue
                fill_w = float(prof['fill_weight_g'])
                rec = form_df[form_df['formula_name'] == fname].iloc[0]['recipe']
                if isinstance(rec, dict):
                    rec_items = [{"Ingredient": k, "%": v} for k, v in rec.items()]
                elif isinstance(rec, list):
                    rec_items = rec
                else:
                    continue
                new_bulk = 0.0
                for rr in rec_items:
                    ing = rr.get('Ingredient'); pct = rr.get('%', 0)
                    req_g = (pct/100) * fill_w
                    m = inv_df[inv_df['trade_name'] == ing]
                    if not m.empty:
                        new_bulk += (req_g/1000) * _rm_avg(m.iloc[0])
                pack_c = float(prof.get('packaging_cost', 0) or 0)
                mfg_c = float(prof.get('mfg_cost', 0) or 0)
                lbl_c = float(prof.get('label_cost', 0) or 0)
                new_total = new_bulk + pack_c + mfg_c + lbl_c
                old_total = float(prof['total_cogs'])
                drift_abs = abs(new_total - old_total)
                drift_pct = (drift_abs / old_total * 100) if old_total > 0 else 0

                if drift_abs <= 0.01:
                    continue  # no meaningful change

                retail = float(prof['target_retail'])
                new_margin = ((retail - new_total) / retail * 100) if retail > 0 else 0

                if drift_pct > 5.0:
                    # Significant drift: archive old version, create new
                    supabase.table('cogs_records').update({"is_active": False}).eq('id', int(prof['id'])).execute()
                    supabase.table('cogs_records').insert({"product_name": prof['product_name'], "formula_name": fname, "fill_weight_g": fill_w, "primary_packaging": prof['primary_packaging'], "bulk_cost": float(new_bulk), "packaging_cost": pack_c, "mfg_cost": mfg_c, "label_cost": lbl_c, "total_cogs": float(new_total), "target_retail": retail, "gross_margin_pct": float(new_margin), "version": int(prof.get('version', 1) or 1) + 1, "is_active": True, "parent_id": int(prof['id'])}).execute()
                else:
                    # Small drift: silent in-place update
                    supabase.table('cogs_records').update({"bulk_cost": float(new_bulk), "total_cogs": float(new_total), "gross_margin_pct": float(new_margin)}).eq('id', int(prof['id'])).execute()
                updates += 1
            if updates > 0:
                _fetch_cached.clear()
        except Exception:
            pass

    # Run auto-sync only once per session reload (not on every minor rerun)
    if "cogs_synced_this_session" not in st.session_state:
        auto_sync_cogs()
        st.session_state.cogs_synced_this_session = True

    user_role = st.session_state.get("user_role", "admin")
    if user_role == "admin":
        MODULES = {
            "📊 Finance & Sales": ["Sales & Revenue", "Analytics", "Clients", "Consignment Tracker", "Financial Overview", "Balance Sheet"],
            "📦 Inventory Management": ["Raw Material Library", "Packaging Library", "Finished Products", "Purchase Requisition"],
            "⚗️ R&D & Production": ["Formula Library", "Formula Builder", "COGS Calculator", "Production Logs"],
            "🛠️ Admin Tools": ["Data Cleaning", "Portfolio Builder", "Price Manager", "Bulk Import"]
        }
    else:
        MODULES = {
            "📦 Modules": ["Formula Library", "Stock Levels"]
        }
    if "active_module" not in st.session_state or st.session_state.active_module not in MODULES:
        st.session_state.active_module = list(MODULES.keys())[0]
        st.session_state.active_nav = MODULES[st.session_state.active_module][0]
    if "active_nav" not in st.session_state: st.session_state.active_nav = MODULES[st.session_state.active_module][0]
    with st.sidebar:
        try: st.image("logo.jpg", use_container_width=True)
        except: st.markdown("<h3 style='text-align: center; padding-bottom: 20px;'>T / O</h3>", unsafe_allow_html=True)
        st.write("##")
        st.markdown("<p style='opacity: 0.6; font-weight: 600; font-size: 0.85rem; text-transform: uppercase;'>Business Module</p>", unsafe_allow_html=True)
        selected_module = st.selectbox("Module", list(MODULES.keys()), index=list(MODULES.keys()).index(st.session_state.active_module), label_visibility="collapsed")
        if selected_module != st.session_state.active_module:
            st.session_state.active_module = selected_module
            st.session_state.active_nav = MODULES[selected_module][0]
            st.rerun()
        st.write("---")
        st.markdown("<p style='opacity: 0.6; font-weight: 600; font-size: 0.85rem; text-transform: uppercase;'>Navigation</p>", unsafe_allow_html=True)
        nav_opts = MODULES[st.session_state.active_module]
        if st.session_state.active_nav not in nav_opts:
            for mod, navs in MODULES.items():
                if st.session_state.active_nav in navs:
                    st.session_state.active_module = mod
                    st.rerun()
        selected_nav = st.radio("Nav", nav_opts, index=nav_opts.index(st.session_state.active_nav) if st.session_state.active_nav in nav_opts else 0, label_visibility="collapsed")
        if selected_nav != st.session_state.active_nav:
            st.session_state.active_nav = selected_nav
            st.rerun()
        menu = st.session_state.active_nav
        st.write("<br><br>", unsafe_allow_html=True)
        st.markdown(f"<p style='opacity: 0.6; font-size: 0.8rem; text-align: center;'>Logged in as {st.session_state.get('user_name', 'User')}</p>", unsafe_allow_html=True)
        if st.button("Log Out", use_container_width=True): st.session_state["authenticated"] = False; st.session_state["user_role"] = None; st.session_state["user_name"] = None; st.rerun()

    # --- 1. SALES & REVENUE ---
    if menu == "Sales & Revenue":
        d = load_tables('sales_records', 'finished_goods', 'packaging', 'clients')
        sales_records_df = d['sales_records']; finished_goods = d['finished_goods']; packaging = d['packaging']; clients_df = d['clients']
        st.title("Sales & Revenue Tracker")
        st.markdown("<p style='opacity: 0.6;'>Monitor order volume, track pending receivables, and manage vault stock deductions.</p>", unsafe_allow_html=True)
        if not sales_records_df.empty:
            sales_records_df['sale_date'] = pd.to_datetime(sales_records_df['sale_date'], errors='coerce')
            sales_records_df['Year'] = sales_records_df['sale_date'].dt.year
            years_available = sorted(sales_records_df['Year'].unique().tolist(), reverse=True)
            c_year, c_target = st.columns([1, 3])
            selected_year = c_year.selectbox("Fiscal Year", years_available)
            annual_target = c_target.number_input("Annual Revenue Target ($)", value=5000, step=5000)
            yr_df = sales_records_df[sales_records_df['Year'] == selected_year]
            yr_rev = yr_df['gross_revenue'].sum()
            yr_profit = yr_df['net_profit'].sum()
            yr_units = yr_df['quantity'].sum()
            avg_margin = (yr_profit / yr_rev * 100) if yr_rev > 0 else 0.0
            global_pending_df = sales_records_df[sales_records_df['status'] == 'Pending'].copy()
            global_pending_rev = global_pending_df['gross_revenue'].sum()
            st.write("---")
            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric(f"{selected_year} Gross Revenue", f"${yr_rev:,.2f}")
            k2.metric("Total Pending Cash", f"${global_pending_rev:,.2f}", delta="-Uncollected" if global_pending_rev > 0 else None, delta_color="inverse")
            k3.metric("Net Profit", f"${yr_profit:,.2f}")
            k4.metric("Avg. Profit Margin", f"{avg_margin:.1f}%")
            k5.metric("Total Units Sold", f"{yr_units:,}")
            progress_pct = min(yr_rev / annual_target, 1.0) if annual_target > 0 else 0.0
            st.write(f"**Annual Target Progress:** {progress_pct*100:.1f}% (${yr_rev:,.0f} / ${annual_target:,.0f})")
            st.progress(progress_pct)
            if not global_pending_df.empty:
                st.write("")
                with st.expander(f"⚠️ View Aging Receivables ({len(global_pending_df)} Unpaid Line Items)"):
                    today = pd.Timestamp(datetime.now().date())
                    global_pending_df['Days Pending'] = (today - global_pending_df['sale_date']).dt.days
                    def format_age(days):
                        if days > 60: return f"🔴 {days} days"
                        elif days > 30: return f"🟠 {days} days"
                        else: return f"🟢 {days} days"
                    global_pending_df['Aging'] = global_pending_df['Days Pending'].apply(format_age)
                    aging_df = global_pending_df.sort_values(by='Days Pending', ascending=False)
                    aging_df['sale_date'] = aging_df['sale_date'].dt.strftime('%Y-%m-%d')
                    st.dataframe(aging_df[['Aging', 'sale_date', 'account', 'order_ref_number', 'order_description', 'gross_revenue']], use_container_width=True, hide_index=True, column_config={"gross_revenue": st.column_config.NumberColumn("Amount Due", format="$%.2f")})
            st.write("---")
            st.markdown("#### Transaction Ledger & Order Management")
            display_sales = yr_df.copy().sort_values('sale_date', ascending=False)
            display_sales['sale_date'] = display_sales['sale_date'].dt.strftime('%Y-%m-%d')
            display_sales.insert(0, '🔍', False)
            with st.container(border=True):
                edited_sales = st.data_editor(
                    display_sales[['🔍', 'id', 'sale_date', 'order_ref_number', 'account', 'order_description', 'quantity', 'gross_revenue', 'net_profit', 'channel', 'status']], 
                    use_container_width=True, hide_index=True,
                    disabled=['id', 'sale_date', 'order_ref_number', 'account', 'order_description', 'quantity', 'gross_revenue', 'net_profit', 'channel'],
                    column_config={
                        "id": None, "gross_revenue": st.column_config.NumberColumn("Gross Rev", format="$%.2f"),
                        "net_profit": st.column_config.NumberColumn("Net Profit", format="$%.2f"),
                        "status": st.column_config.SelectboxColumn("Status", options=["Paid", "Pending", "Cancelled"], required=True)
                    }
                )
                if st.button("💾 Synchronize Ledger Status", type="primary"):
                    for index, row in edited_sales.iterrows():
                        orig = display_sales.loc[index]
                        if row['status'] != orig['status']:
                            supabase.table('sales_records').update({'status': row['status']}).eq('id', int(row['id'])).execute()
                    st.success("Ledger payments synchronized!")
                    clear_cache(); st.rerun()
            selected_sales = edited_sales[edited_sales['🔍'] == True]
            if not selected_sales.empty:
                sel_id = selected_sales.iloc[0]['id']
                sale_item = yr_df[yr_df['id'] == sel_id].iloc[0]
                ref_num = sale_item['order_ref_number']
                if pd.notna(ref_num) and str(ref_num).strip() != "":
                    order_items = yr_df[yr_df['order_ref_number'] == ref_num]
                else:
                    order_items = pd.DataFrame([sale_item])
                st.write("##")
                with st.container(border=True):
                    display_ref = f"ORD-{(int(ref_num) + 200):06d}" if str(ref_num).isdigit() else ref_num
                    st.markdown(f"#### 📦 Inspecting Order Reference: {display_ref if pd.notna(ref_num) else 'Unreferenced'}")
                    st.write(f"**Client:** {sale_item['account']} | **Date:** {sale_item['sale_date'].strftime('%Y-%m-%d')}")
                    st.dataframe(order_items[['order_description', 'quantity', 'unit_price', 'gross_revenue']], hide_index=True, use_container_width=True)
                    order_total = order_items['gross_revenue'].sum()
                    st.metric("Total Order Value", f"${order_total:,.2f}")
                    with st.expander("✏️ Edit Order Pricing & Apply Discounts"):
                        edit_rows = []
                        for _, orow in order_items.iterrows():
                            edit_rows.append({"id": int(orow['id']), "Product": orow['order_description'], "Qty": int(orow['quantity']), "Unit Price": float(orow['unit_price']), "Disc %": 0.0, "Current Revenue": float(orow['gross_revenue'])})
                        edit_order_df = pd.DataFrame(edit_rows)
                        edited_order = st.data_editor(edit_order_df, use_container_width=True, hide_index=True, disabled=['id', 'Product', 'Qty', 'Current Revenue'], column_config={"id": None, "Current Revenue": st.column_config.NumberColumn(format="$%.2f"), "Unit Price": st.column_config.NumberColumn(format="$%.2f"), "Disc %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=5.0)})
                        flat_disc = st.number_input("Order-Level Flat Discount ($)", min_value=0.0, value=0.0, step=1.0, key="edit_ord_disc")
                        order_note = st.text_input("Order Note / Comment", placeholder="e.g., 10% loyalty discount applied", key="edit_ord_note")
                        new_subtotal = sum(r['Qty'] * r['Unit Price'] * (1 - r['Disc %'] / 100) for _, r in edited_order.iterrows())
                        new_total = max(0, new_subtotal - flat_disc)
                        st.markdown(f"**New Order Total: ${new_total:,.2f}**")
                        if st.button("💾 Apply Changes to This Order", type="primary"):
                            num_lines = len(edited_order)
                            flat_per_line = flat_disc / num_lines if num_lines > 0 else 0
                            for _, erow in edited_order.iterrows():
                                line_gross = erow['Qty'] * erow['Unit Price'] * (1 - erow['Disc %'] / 100) - flat_per_line
                                line_gross = max(0, line_gross)
                                # Recalculate net profit using original COGS
                                orig_row = order_items[order_items['id'] == erow['id']].iloc[0]
                                orig_cogs = float(orig_row['cogs']) if 'cogs' in orig_row and pd.notna(orig_row['cogs']) else 0.0
                                line_net = line_gross - orig_cogs
                                line_gm = (line_net / line_gross) if line_gross > 0 else 0.0
                                update_data = {"unit_price": float(erow['Unit Price']), "gross_revenue": float(line_gross), "net_profit": float(line_net), "gm": float(line_gm)}
                                disc_parts = []
                                if erow['Disc %'] > 0: disc_parts.append(f"{erow['Disc %']:.0f}% line discount")
                                if flat_disc > 0: disc_parts.append(f"${flat_disc:.2f} order discount")
                                note_text = order_note if order_note else ""
                                if disc_parts: note_text = (note_text + " | " if note_text else "") + ", ".join(disc_parts)
                                if note_text: update_data["notes"] = note_text
                                supabase.table('sales_records').update(update_data).eq('id', int(erow['id'])).execute()
                            st.success("Order updated with new pricing!")
                            time.sleep(1)
                            clear_cache(); st.rerun()
                    col_pdf, col_rev = st.columns(2)
                    with col_pdf:
                        pdf_bytes = generate_order_pdf(str(ref_num), order_items, str(sale_item['account']), sale_item['sale_date'].strftime('%Y-%m-%d'))
                        file_ref = f"ORD-{(int(ref_num) + 200):06d}" if str(ref_num).isdigit() else ref_num
                        st.download_button(label="📄 Download PDF Order Summary", data=pdf_bytes, file_name=f"TherapeuticOils_Order_{file_ref}.pdf", mime="application/pdf", use_container_width=True)
                    with col_rev:
                        with st.expander("⚠️ System Actions: Reverse Line Item"):
                            rev_pass = st.text_input("Authorization Passcode", type="password", key=f"rev_{sel_id}")
                            if st.button("Reverse Sale & Restore Stock", type="primary"):
                                if rev_pass == "lab2026":
                                    fp_match = finished_goods[finished_goods['product_name'] == sale_item['order_description']]
                                    if not fp_match.empty:
                                        fp_id = int(fp_match.iloc[0]['id'])
                                        current_stock = int(fp_match.iloc[0]['stock_quantity'])
                                        new_stock = current_stock + int(sale_item['quantity'])
                                        supabase.table('finished_products').update({'stock_quantity': new_stock}).eq('id', fp_id).execute()
                                    supabase.table('sales_records').delete().eq('id', int(sel_id)).execute()
                                    st.success("Transaction reversed! Financials updated and FP stock restored.")
                                    time.sleep(1)
                                    clear_cache(); st.rerun()
                                else:
                                    st.error("Incorrect passcode.")
        else:
            st.info("No sales records imported or logged yet.")
        st.write("---")
        with st.expander("➕ Log New Sales Order", expanded=False):
            if not finished_goods.empty:
                pkg_opts = ["None"]
                if not packaging.empty: pkg_opts += packaging['material_name'].tolist()
                next_ord_id = 200
                if not sales_records_df.empty:
                    refs = sales_records_df['order_ref_number'].astype(str)
                    new_fmt = refs.str.extract(r'ORD-(\d+)')[0].dropna().astype(int)
                    old_fmt = refs[refs.str.match(r'^\d+$')].astype(int) + 200
                    all_ids = pd.concat([new_fmt, old_fmt]).dropna()
                    if not all_ids.empty:
                        next_ord_id = max(200, int(all_ids.max()) + 1)
                default_ord_ref = f"ORD-{next_ord_id:06d}"
                client_opts = ["-- Type manually --"]
                if not clients_df.empty:
                    client_opts += clients_df['client_name'].tolist()
                st.markdown("#### 1. Client & Order Info")
                h1, h2, h3 = st.columns(3)
                client_select = h1.selectbox("Select Client", client_opts)
                if client_select == "-- Type manually --":
                    client_name = h2.text_input("Client Name", placeholder="e.g., Ralph J. Ghosn")
                    client_channel = h3.selectbox("Channel", ["Physiotherapists", "Beauty centers", "Direct to Consumer", "Wholesale"])
                else:
                    client_name = client_select
                    matched_client = clients_df[clients_df['client_name'] == client_select].iloc[0]
                    client_channel = h3.text_input("Channel", value=str(matched_client['channel']), disabled=True)
                    h2.text_input("Client Name", value=client_name, disabled=True)
                h4, h5, h6 = st.columns(3)
                order_ref = h4.text_input("Order Ref #", value=default_ord_ref)
                sale_date = h5.date_input("Date of Sale", value=datetime.today())
                status = h6.selectbox("Payment Status", ["Paid", "Pending", "Cancelled"])
                st.write("---")
                st.markdown("#### 2. Order Line Items")
                st.info("💡 Add multiple products to this order. Each row is one line item.")
                fp_opts = finished_goods['product_name'].tolist()
                if "order_lines" not in st.session_state:
                    st.session_state.order_lines = [{"product": fp_opts[0], "qty": 1, "price": None}]
                lines_to_remove = None
                for i, line in enumerate(st.session_state.order_lines):
                    lc1, lc2, lc3, lc4, lc5 = st.columns([3, 1, 1, 0.8, 0.5])
                    line['product'] = lc1.selectbox("Product", fp_opts, index=fp_opts.index(line['product']) if line['product'] in fp_opts else 0, key=f"ol_prod_{i}")
                    line['qty'] = lc2.number_input("Qty", min_value=1, value=line['qty'], step=1, key=f"ol_qty_{i}")
                    fg_m = finished_goods[finished_goods['product_name'] == line['product']].iloc[0]
                    default_p = float(fg_m['retail_price'])
                    line['price'] = lc3.number_input("Unit $", min_value=0.0, value=line['price'] if line['price'] is not None else default_p, step=0.5, key=f"ol_price_{i}")
                    if 'disc' not in line: line['disc'] = 0.0
                    line['disc'] = lc4.number_input("Disc %", min_value=0.0, max_value=100.0, value=line['disc'], step=5.0, key=f"ol_disc_{i}")
                    if i > 0:
                        if lc5.button("✕", key=f"ol_del_{i}"):
                            lines_to_remove = i
                if lines_to_remove is not None:
                    st.session_state.order_lines.pop(lines_to_remove)
                    clear_cache(); st.rerun()
                if st.button("＋ Add Another Product"):
                    st.session_state.order_lines.append({"product": fp_opts[0], "qty": 1, "price": None})
                    clear_cache(); st.rerun()
                st.write("---")
                st.markdown("#### 3. Order Discount")
                ord_disc = st.number_input("Order-Level Discount ($)", min_value=0.0, value=0.0, step=1.0, key="ord_disc")
                preview_subtotal = sum(l['qty'] * (l['price'] or 0) * (1 - l.get('disc', 0) / 100) for l in st.session_state.order_lines)
                preview_total = max(0, preview_subtotal - ord_disc)
                preview_savings = sum(l['qty'] * (l['price'] or 0) * (l.get('disc', 0) / 100) for l in st.session_state.order_lines) + ord_disc
                st.markdown(f"**Subtotal (after line discounts): ${preview_subtotal:,.2f}**")
                if preview_savings > 0:
                    st.markdown(f"**Total Savings: -${preview_savings:,.2f}**")
                st.markdown(f"### Order Total: ${preview_total:,.2f} · {len(st.session_state.order_lines)} line item(s)")
                st.write("---")
                st.markdown("#### 4. Fulfillment Materials")
                default_f_df = pd.DataFrame([{"Fulfillment Material": "None", "Quantity": 1}])
                f_edited = st.data_editor(default_f_df, num_rows="dynamic", use_container_width=True, hide_index=True, key="multiline_fulfill", column_config={"Fulfillment Material": st.column_config.SelectboxColumn("Fulfillment Material", options=pkg_opts, required=True), "Quantity": st.column_config.NumberColumn("Quantity", min_value=1, step=1, required=True)})
                st.write("---")
                if st.button("🚀 Submit Entire Order & Deduct Stock", type="primary", use_container_width=True):
                    if not client_name:
                        st.error("⚠️ Please select or enter a client name.")
                    else:
                        shortage = False
                        for line in st.session_state.order_lines:
                            fg_m = finished_goods[finished_goods['product_name'] == line['product']].iloc[0]
                            if int(fg_m['stock_quantity']) < line['qty']:
                                st.error(f"⚠️ Not enough {line['product']} in stock ({fg_m['stock_quantity']} available, {line['qty']} needed).")
                                shortage = True
                                break
                        f_needs = {}
                        for _, f_row in f_edited.iterrows():
                            item = f_row.get("Fulfillment Material")
                            q = f_row.get("Quantity")
                            if pd.notna(item) and item != "None" and pd.notna(q):
                                f_needs[item] = f_needs.get(item, 0) + int(q)
                        fulfillment_cost = 0.0
                        pkg_updates = []
                        for item, q in f_needs.items():
                            pm_match = packaging[packaging['material_name'] == item]
                            if not pm_match.empty:
                                pm_id = int(pm_match.iloc[0]['id'])
                                pm_cost = float(pm_match.iloc[0]['cost_per_unit'])
                                pm_stock = int(pm_match.iloc[0]['remaining_quantity'])
                                if pm_stock < q:
                                    shortage = True
                                    st.error(f"⚠️ Not enough '{item}' in Packaging Vault.")
                                    break
                                fulfillment_cost += (pm_cost * q)
                                pkg_updates.append({"id": pm_id, "new_stock": pm_stock - q})
                        if not shortage:
                            total_units = sum(l['qty'] for l in st.session_state.order_lines)
                            for line in st.session_state.order_lines:
                                fg_m = finished_goods[finished_goods['product_name'] == line['product']].iloc[0]
                                unit_cogs = float(fg_m['unit_cogs'])
                                gross = line['qty'] * line['price']
                                line_fulfill = fulfillment_cost * (line['qty'] / total_units) if total_units > 0 else 0
                                total_cogs = (line['qty'] * unit_cogs) + line_fulfill
                                net = gross - total_cogs
                                gm = (net / gross) if gross > 0 else 0.0
                                new_stock = int(fg_m['stock_quantity']) - line['qty']
                                supabase.table('finished_products').update({'stock_quantity': new_stock}).eq('id', int(fg_m['id'])).execute()
                                supabase.table('sales_records').insert({"order_description": line['product'], "quantity": line['qty'], "unit_price": line['price'], "gross_revenue": gross, "cogs": total_cogs, "net_profit": net, "account": client_name, "order_ref_number": order_ref, "sale_date": sale_date.strftime('%Y-%m-%d'), "gm": gm, "channel": client_channel, "status": status}).execute()
                            for pu in pkg_updates:
                                supabase.table('packaging').update({'remaining_quantity': pu['new_stock']}).eq('id', pu['id']).execute()
                            if client_select == "-- Type manually --" and client_name:
                                if clients_df.empty or client_name not in clients_df['client_name'].values:
                                    supabase.table('clients').insert({"client_name": client_name, "channel": client_channel}).execute()
                            st.session_state.order_lines = [{"product": fp_opts[0], "qty": 1, "price": None}]
                            st.success(f"✅ Order {order_ref} logged! All stock deducted.")
                            time.sleep(1.5)
                            clear_cache(); st.rerun()

    # --- 1.1 CLIENTS DATABASE ---
    elif menu == "Clients":
        d = load_tables('clients', 'sales_records')
        clients_df = d['clients']; sales_records_df = d['sales_records']
        st.title("Client Database")
        st.markdown("<p style='opacity: 0.6;'>Manage your client and account records. Select a client to view their order history.</p>", unsafe_allow_html=True)
        if not clients_df.empty:
            display_clients = clients_df.copy()
            display_clients.insert(0, '🔍', False)
            with st.container(border=True):
                edited_clients = st.data_editor(
                    display_clients[['🔍', 'id', 'client_name', 'business_name', 'phone', 'email', 'channel']],
                    use_container_width=True, hide_index=True, disabled=['id'],
                    column_config={"id": None, "channel": st.column_config.SelectboxColumn("Channel", options=["Physiotherapists", "Beauty centers", "Direct to Consumer", "Wholesale"], required=True)}
                )
                if st.button("💾 Synchronize Client Records", type="primary"):
                    for idx, row in edited_clients.iterrows():
                        orig = clients_df.loc[idx]
                        if any(row[c] != orig[c] for c in ['client_name', 'business_name', 'phone', 'email', 'channel']):
                            supabase.table('clients').update({"client_name": row['client_name'], "business_name": row['business_name'], "phone": row['phone'], "email": row['email'], "channel": row['channel']}).eq('id', int(orig['id'])).execute()
                    st.success("Client records synchronized!")
                    clear_cache(); st.rerun()
            selected_clients = edited_clients[edited_clients['🔍'] == True]
            if not selected_clients.empty:
                client_row = clients_df.loc[selected_clients.index[0]]
                st.write("##")
                with st.container(border=True):
                    st.markdown(f"#### {client_row['client_name']}")
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Business:** {client_row['business_name'] or 'N/A'}<br>**Channel:** {client_row['channel']}", unsafe_allow_html=True)
                    c2.write(f"**Phone:** {client_row['phone'] or 'N/A'}<br>**Email:** {client_row['email'] or 'N/A'}", unsafe_allow_html=True)
                    c3.write(f"**Address:** {client_row['address'] or 'N/A'}")
                    if client_row.get('notes'):
                        st.write(f"**Notes:** {client_row['notes']}")
                    st.write("---")
                    st.markdown("#### 📋 Order History")
                    if not sales_records_df.empty:
                        client_sales = sales_records_df[sales_records_df['account'] == client_row['client_name']].copy()
                        if not client_sales.empty:
                            client_sales['sale_date'] = pd.to_datetime(client_sales['sale_date'], errors='coerce').dt.strftime('%Y-%m-%d')
                            st.dataframe(client_sales[['sale_date', 'order_ref_number', 'order_description', 'quantity', 'gross_revenue', 'status']].sort_values('sale_date', ascending=False), use_container_width=True, hide_index=True, column_config={"gross_revenue": st.column_config.NumberColumn("Revenue", format="$%.2f")})
                            total_rev = client_sales['gross_revenue'].sum()
                            total_orders = client_sales['order_ref_number'].nunique()
                            r1, r2 = st.columns(2)
                            r1.metric("Lifetime Revenue", f"${total_rev:,.2f}")
                            r2.metric("Total Orders", f"{total_orders}")
                        else:
                            st.info("No orders found for this client.")
                    else:
                        st.info("No sales records in the system yet.")
                    with st.expander("System Actions"):
                        del_pass = st.text_input("Authorization Passcode", type="password", key="del_client")
                        if st.button("Erase Client Record") and del_pass == "lab2026":
                            supabase.table('clients').delete().eq('id', int(client_row['id'])).execute()
                            clear_cache(); st.rerun()
        else:
            st.info("No clients registered yet.")
        st.write("---")
        with st.expander("➕ Register New Client"):
            with st.form("add_client", clear_on_submit=True):
                ac1, ac2 = st.columns(2)
                new_name = ac1.text_input("Client Name *")
                new_biz = ac2.text_input("Clinic / Business Name")
                ac3, ac4 = st.columns(2)
                new_phone = ac3.text_input("Phone")
                new_email = ac4.text_input("Email")
                new_addr = st.text_input("Address")
                ac5, ac6 = st.columns(2)
                new_channel = ac5.selectbox("Channel", ["Physiotherapists", "Beauty centers", "Direct to Consumer", "Wholesale"])
                new_notes = ac6.text_input("Notes")
                if st.form_submit_button("Register Client", type="primary") and new_name:
                    supabase.table('clients').insert({"client_name": new_name, "business_name": new_biz, "phone": new_phone, "email": new_email, "address": new_addr, "channel": new_channel, "notes": new_notes}).execute()
                    st.success(f"Client '{new_name}' registered!")
                    time.sleep(1); clear_cache(); st.rerun()

    # --- 1.5 CONSIGNMENT TRACKER ---
    elif menu == "Consignment Tracker":
        d = load_tables('consignment', 'finished_goods')
        consignment_df = d['consignment']; finished_goods = d['finished_goods']
        st.title("Consignment Agreements")
        st.markdown("<p style='opacity: 0.6;'>Manage goods sitting on partner shelves. Consigned goods are deducted from your lab stock but do not count as Revenue until explicitly marked as sold here.</p>", unsafe_allow_html=True)
        if not consignment_df.empty:
            active_cons = consignment_df[consignment_df['status'] == 'Active'].copy()
            total_active_units = active_cons['qty_consigned'].sum() - active_cons['qty_sold'].sum()
            total_potential_rev = ((active_cons['qty_consigned'] - active_cons['qty_sold']) * active_cons['wholesale_price']).sum()
            col1, col2 = st.columns(2)
            col1.metric("Unsold Units on Partner Shelves", f"{total_active_units:,}")
            col2.metric("Total Potential Payout Revenue", f"${total_potential_rev:,.2f}")
            if not active_cons.empty:
                with st.expander("📄 Generate Partner Inventory Statement"):
                    partner_names = active_cons['partner_name'].unique().tolist()
                    sel_partner = st.selectbox("Select Partner", partner_names, key="partner_inv_select")
                    partner_items = active_cons[active_cons['partner_name'] == sel_partner].copy()
                    partner_items['Remaining'] = partner_items['qty_consigned'] - partner_items['qty_sold']
                    partner_items = partner_items[partner_items['Remaining'] > 0]
                    if not partner_items.empty:
                        st.dataframe(partner_items[['product_name', 'order_ref_number', 'qty_consigned', 'qty_sold', 'Remaining', 'retail_price', 'wholesale_price']].rename(columns={'product_name': 'Product', 'order_ref_number': 'Ref #', 'qty_consigned': 'Sent', 'qty_sold': 'Sold', 'retail_price': 'Retail $', 'wholesale_price': 'Owed/Unit $'}), use_container_width=True, hide_index=True)
                        inv_pdf = generate_partner_inventory_pdf(sel_partner, partner_items, datetime.today().strftime('%Y-%m-%d'))
                        st.download_button(label=f"📄 Download {sel_partner} Inventory PDF", data=inv_pdf, file_name=f"PartnerInventory_{sel_partner.replace(' ', '_')}_{datetime.today().strftime('%Y%m%d')}.pdf", mime="application/pdf", use_container_width=True, type="primary")
                    else:
                        st.info(f"No remaining unsold items for {sel_partner}.")
            st.write("---")
            st.markdown("#### Active Consignment Ledgers")
            display_cons = consignment_df.copy().sort_values('created_at', ascending=False)
            display_cons['Date'] = pd.to_datetime(display_cons['created_at'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('N/A')
            display_cons['Remaining'] = display_cons['qty_consigned'] - display_cons['qty_sold']
            display_cons.insert(0, '🔍', False)
            with st.container(border=True):
                edited_cons = st.data_editor(display_cons[['🔍', 'id', 'Date', 'partner_name', 'order_ref_number', 'product_name', 'qty_consigned', 'qty_sold', 'Remaining', 'wholesale_price', 'retail_price', 'status']], use_container_width=True, hide_index=True, disabled=['id', 'Date', 'partner_name', 'order_ref_number', 'product_name', 'qty_consigned', 'qty_sold', 'Remaining', 'wholesale_price', 'retail_price', 'status'], column_config={"id": None, "wholesale_price": st.column_config.NumberColumn("Your Price", format="$%.2f"), "retail_price": st.column_config.NumberColumn("Retail Price", format="$%.2f")})
            selected_cons = edited_cons[edited_cons['🔍'] == True]
            if not selected_cons.empty:
                sel_id = selected_cons.iloc[0]['id']
                cons_item = consignment_df[consignment_df['id'] == sel_id].iloc[0]
                ref_num = cons_item['order_ref_number']
                if pd.notna(ref_num) and str(ref_num).strip() != "":
                    batch_items = consignment_df[consignment_df['order_ref_number'] == ref_num]
                else:
                    batch_items = pd.DataFrame([cons_item])
                st.write("##")
                with st.container(border=True):
                    st.markdown(f"#### 🤝 Inspecting Consignment: {ref_num if pd.notna(ref_num) else 'Unreferenced'}")
                    st.write(f"**Partner:** {cons_item['partner_name']}")
                    pdf_bytes = generate_consignment_pdf(str(ref_num), batch_items, str(cons_item['partner_name']), pd.to_datetime(cons_item['created_at'], errors='coerce').strftime('%Y-%m-%d'))
                    st.download_button(label="📄 Download Official Consignment Agreement PDF", data=pdf_bytes, file_name=f"Consignment_{ref_num}.pdf", mime="application/pdf", use_container_width=True, type="secondary")
                    st.write("---")
                    st.markdown(f"**Log Sales for: {cons_item['product_name']}**")
                    remaining_to_sell = int(cons_item['qty_consigned']) - int(cons_item['qty_sold'])
                    if remaining_to_sell > 0:
                        with st.form("log_cons_sale"):
                            c1, c2 = st.columns(2)
                            units_sold = c1.number_input("Units Sold by Partner", min_value=1, max_value=int(remaining_to_sell), step=1)
                            payment_status = c2.selectbox("Has the partner paid you for these yet?", ["Pending", "Paid"])
                            if st.form_submit_button("Log as Revenue & Update Consignment", type="primary"):
                                new_qty_sold = int(cons_item['qty_sold']) + int(units_sold)
                                new_status = "Completed" if new_qty_sold >= int(cons_item['qty_consigned']) else "Active"
                                supabase.table('consignment_records').update({'qty_sold': int(new_qty_sold), 'status': new_status}).eq('id', int(sel_id)).execute()
                                gross_rev = float(units_sold) * float(cons_item['wholesale_price'])
                                cogs = float(units_sold) * float(cons_item['unit_cogs'])
                                net_profit = float(gross_rev - cogs)
                                gm = float(net_profit / gross_rev) if gross_rev > 0 else 0.0
                                supabase.table('sales_records').insert({"order_description": str(cons_item['product_name']), "quantity": int(units_sold), "unit_price": float(cons_item['wholesale_price']), "gross_revenue": float(gross_rev), "cogs": float(cogs), "net_profit": float(net_profit), "account": str(cons_item['partner_name']), "order_ref_number": str(ref_num), "sale_date": datetime.today().strftime('%Y-%m-%d'), "gm": float(gm), "channel": "Consignment Payout", "status": payment_status}).execute()
                                st.success(f"Successfully converted {units_sold} consigned units into Sales Revenue!")
                                time.sleep(1.5); clear_cache(); st.rerun()
                    else:
                        st.success("✅ All units from this consignment line have been sold and logged.")
                if st.session_state.get("user_role") == "admin":
                    st.write("---")
                    st.markdown("**⚠️ Admin: Reset Consignment**")
                    reset_col1, reset_col2 = st.columns([2, 1])
                    reset_qty = reset_col1.number_input("Set qty_sold to:", min_value=0, max_value=int(cons_item['qty_consigned']), value=int(cons_item['qty_sold']), step=1, key="reset_cons_qty")
                    reset_col2.write("<br>", unsafe_allow_html=True)
                    if reset_col2.button("Reset", type="primary", key="reset_cons_btn"):
                        new_s = "Active" if int(reset_qty) < int(cons_item['qty_consigned']) else "Completed"
                        supabase.table('consignment_records').update({'qty_sold': int(reset_qty), 'status': new_s}).eq('id', int(sel_id)).execute()
                        _fetch_cached.clear()
                        st.cache_data.clear()
                        for key in list(st.session_state.keys()):
                            if key not in ['authenticated', 'user_role', 'user_name', 'active_module', 'active_nav']:
                                del st.session_state[key]
                        st.rerun()
                    with st.expander("🗑️ Delete Consignment Record"):
                        del_cons_pass = st.text_input("Authorization Passcode", type="password", key="del_cons_pass")
                        restore_stock = st.checkbox("Restore stock to Finished Products", value=True, key="restore_cons_stock")
                        if st.button("Permanently Delete This Record", key="del_cons_btn"):
                            if del_cons_pass == "lab2026":
                                if restore_stock:
                                    remaining = int(cons_item['qty_consigned']) - int(cons_item['qty_sold'])
                                    fp_match = finished_goods[finished_goods['product_name'] == cons_item['product_name']]
                                    if not fp_match.empty:
                                        fp_id = int(fp_match.iloc[0]['id'])
                                        current_stock = int(fp_match.iloc[0]['stock_quantity'])
                                        supabase.table('finished_products').update({'stock_quantity': current_stock + remaining}).eq('id', fp_id).execute()
                                supabase.table('consignment_records').delete().eq('id', int(sel_id)).execute()
                                _fetch_cached.clear()
                                st.cache_data.clear()
                                for key in list(st.session_state.keys()):
                                    if key not in ['authenticated', 'user_role', 'user_name', 'active_module', 'active_nav']:
                                        del st.session_state[key]
                                st.rerun()
                            else:
                                st.error("Incorrect passcode.")
        else:
            st.info("No consignment records found.")
        st.write("---")
        with st.expander("➕ Consign New Goods (Deducts from Lab Stock)"):
            if not finished_goods.empty:
                next_cons_id = 250
                if not consignment_df.empty:
                    cons_codes = consignment_df['order_ref_number'].astype(str).str.extract(r'CONS-(\d+)')[0].dropna().astype(int)
                    if not cons_codes.empty:
                        next_cons_id = max(250, cons_codes.max() + 1)
                default_ref = f"CONS-{next_cons_id:06d}"
                with st.form("add_consignment"):
                    st.info("💡 Goods entered here will leave your inventory vault but will NOT count towards Gross Revenue until the partner sells them.")
                    c1, c2, c3 = st.columns(3)
                    partner = c1.text_input("Partner / Retailer Name")
                    ref = c2.text_input("Consignment Ref #", value=default_ref)
                    prod = c3.selectbox("Finished Product", finished_goods['product_name'].tolist())
                    fg_match = finished_goods[finished_goods['product_name'] == prod].iloc[0]
                    def_retail = float(fg_match['retail_price'])
                    def_cogs = float(fg_match['unit_cogs'])
                    curr_stock = int(fg_match['stock_quantity'])
                    c4, c5 = st.columns(2)
                    qty = c4.number_input("Qty to Consign", min_value=1, step=1)
                    retail_p = c5.number_input("Retail Price per Unit (partner sells at) ($)", value=def_retail, min_value=0.0)
                    c6, c7 = st.columns(2)
                    wholesale_p = c6.number_input("Your Price per Unit (partner pays you) ($)", value=def_retail * 0.5, min_value=0.0)
                    c7.write("")
                    c7.write(f"**Partner Margin:** ${retail_p - wholesale_p:.2f}/unit ({((retail_p - wholesale_p) / retail_p * 100) if retail_p > 0 else 0:.0f}%)")
                    if st.form_submit_button("Ship Consignment & Deduct Stock", type="primary"):
                        if not partner or not ref:
                            st.error("⚠️ Please provide both the Partner Name and Consignment Ref #.")
                        elif curr_stock < qty:
                            st.error(f"⚠️ You only have {curr_stock} of {prod}. Aborted.")
                        else:
                            supabase.table('finished_products').update({'stock_quantity': curr_stock - qty}).eq('id', int(fg_match['id'])).execute()
                            supabase.table('consignment_records').insert({"partner_name": str(partner), "order_ref_number": str(ref), "product_name": str(prod), "qty_consigned": int(qty), "unit_cogs": float(def_cogs), "retail_price": float(retail_p), "wholesale_price": float(wholesale_p)}).execute()
                            st.success("Consignment logged securely!")
                            time.sleep(1.5); clear_cache(); st.rerun()

    # --- 2. FINANCIAL OVERVIEW ---
    elif menu == "Financial Overview":
        d = load_tables('inventory', 'packaging', 'finished_goods', 'consignment')
        inventory = d['inventory']; packaging = d['packaging']; finished_goods = d['finished_goods']; consignment_df = d['consignment']
        st.title("Financial Overview")
        st.markdown("<p style='opacity: 0.6;'>Live tracking of physical assets, inventory valuation, and retail projections.</p>", unsafe_allow_html=True)
        st.write("##")
        rm_total = (inventory['price_per_kg'] * inventory['quantity_kg']).sum() if not inventory.empty else 0.0
        pm_total = (packaging['cost_per_unit'] * packaging['remaining_quantity']).sum() if not packaging.empty else 0.0
        fp_cogs_total = (finished_goods['unit_cogs'] * finished_goods['stock_quantity']).sum() if not finished_goods.empty else 0.0
        fp_retail_total = (finished_goods['retail_price'] * finished_goods['stock_quantity']).sum() if not finished_goods.empty else 0.0
        cons_cogs_total = 0.0
        if not consignment_df.empty:
            active_cons = consignment_df[consignment_df['status'] == 'Active'].copy()
            active_cons['unsold_qty'] = active_cons['qty_consigned'] - active_cons['qty_sold']
            cons_cogs_total = (active_cons['unsold_qty'] * active_cons['unit_cogs']).sum()
        vault_assets = rm_total + pm_total + fp_cogs_total + cons_cogs_total
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("Raw Materials", f"${rm_total:,.2f}")
        with c2: st.metric("Packaging", f"${pm_total:,.2f}")
        with c3: st.metric("Finished Goods (In Lab)", f"${fp_cogs_total:,.2f}")
        with c4: st.metric("Finished Goods (Consigned)", f"${cons_cogs_total:,.2f}")
        st.write("---")
        st.markdown("#### Projected Revenue")
        st.metric("Potential Retail Value on Shelf", f"${fp_retail_total:,.2f}", f"Est. Gross Profit: ${(fp_retail_total - fp_cogs_total):,.2f}")

    # --- 3. BALANCE SHEET GENERATOR ---
    elif menu == "Balance Sheet":
        d = load_tables('inventory', 'packaging', 'finished_goods', 'consignment', 'sales_records')
        inventory = d['inventory']; packaging = d['packaging']; finished_goods = d['finished_goods']; consignment_df = d['consignment']; sales_records_df = d['sales_records']
        st.title("Balance Sheet Generator")
        st.markdown("<p style='opacity: 0.6;'>Generate a professional financial statement summarizing assets, liabilities, and owner's equity.</p>", unsafe_allow_html=True)
        rm_total = (inventory['price_per_kg'] * inventory['quantity_kg']).sum() if not inventory.empty else 0.0
        pm_total = (packaging['cost_per_unit'] * packaging['remaining_quantity']).sum() if not packaging.empty else 0.0
        fp_cogs_total = (finished_goods['unit_cogs'] * finished_goods['stock_quantity']).sum() if not finished_goods.empty else 0.0
        cons_cogs_total = 0.0
        if not consignment_df.empty:
            active_cons = consignment_df[consignment_df['status'] == 'Active'].copy()
            active_cons['unsold_qty'] = active_cons['qty_consigned'] - active_cons['qty_sold']
            cons_cogs_total = (active_cons['unsold_qty'] * active_cons['unit_cogs']).sum()
        total_inv_fg = fp_cogs_total + cons_cogs_total
        ar_total = 0.0
        if not sales_records_df.empty:
            ar_total = sales_records_df[sales_records_df['status'] == 'Pending']['gross_revenue'].sum()
        with st.form("balance_sheet_form"):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### ASSETS")
                cash = st.number_input("Cash in Bank ($)", min_value=0.0, value=0.0, step=100.0)
                st.write(f"**Accounts Receivable (Pending Sales):** ${ar_total:,.2f}")
                st.write(f"**Inventory (Raw Materials):** ${rm_total:,.2f}")
                st.write(f"**Inventory (Packaging):** ${pm_total:,.2f}")
                st.write(f"**Inventory (Finished Goods + Consigned):** ${total_inv_fg:,.2f}")
                st.write("---")
                fixed_assets = st.number_input("Property & Equipment Value ($)", min_value=0.0, value=0.0, step=100.0)
            with col2:
                st.markdown("#### LIABILITIES")
                accounts_payable = st.number_input("Accounts Payable (Unpaid Bills) ($)", min_value=0.0, value=0.0, step=100.0)
                debt = st.number_input("Short/Long Term Debt ($)", min_value=0.0, value=0.0, step=100.0)
            submit_bs = st.form_submit_button("Calculate & Generate Balance Sheet", type="primary", use_container_width=True)
        if submit_bs:
            total_assets = cash + ar_total + rm_total + pm_total + total_inv_fg + fixed_assets
            total_liabilities = accounts_payable + debt
            owner_equity = total_assets - total_liabilities
            r1, r2, r3 = st.columns(3)
            r1.metric("Total Assets", f"${total_assets:,.2f}")
            r2.metric("Total Liabilities", f"${total_liabilities:,.2f}")
            r3.metric("Owner's Equity", f"${owner_equity:,.2f}")
            if owner_equity == (total_assets - total_liabilities):
                date_str = datetime.today().strftime('%B %d, %Y')
                pdf_bytes = generate_balance_sheet_pdf(date_str, cash, ar_total, rm_total, pm_total, total_inv_fg, fixed_assets, accounts_payable, debt, total_assets, total_liabilities, owner_equity)
                st.download_button("📄 Download Official PDF Balance Sheet", data=pdf_bytes, file_name=f"TherapeuticOils_BalanceSheet_{datetime.today().strftime('%Y%m%d')}.pdf", mime="application/pdf", use_container_width=True)

    # --- ANALYTICS DASHBOARD ---
    elif menu == "Analytics":
        d = load_tables('sales_records', 'cogs_records', 'finished_goods')
        sales_records_df = d['sales_records']; cogs_records_df = d['cogs_records']; finished_goods = d['finished_goods']
        portfolios_df = fetch_vault_data('portfolios', 'portfolio_name')
        st.title("Analytics Dashboard")
        st.markdown("<p style='opacity: 0.6;'>Visual insights into revenue, product performance, and profitability trends.</p>", unsafe_allow_html=True)
        if not sales_records_df.empty:
            sales_records_df['sale_date'] = pd.to_datetime(sales_records_df['sale_date'], errors='coerce')
            sales_records_df['Year'] = sales_records_df['sale_date'].dt.year
            sales_records_df['Month'] = sales_records_df['sale_date'].dt.to_period('M').astype(str)
            sales_records_df['MonthNum'] = sales_records_df['sale_date'].dt.month
            all_skus = sorted(sales_records_df['order_description'].dropna().unique().tolist())
            years_available = sorted(sales_records_df['Year'].dropna().unique().tolist(), reverse=True)

            # --- Filters ---
            st.write("---")
            fc1, fc2, fc3 = st.columns(3)
            selected_years = fc1.multiselect("Filter by Year", years_available, default=years_available[:2] if len(years_available) >= 2 else years_available)
            selected_skus = fc2.multiselect("Filter by Product (leave empty for all)", all_skus)
            # Portfolio filter
            portfolio_opts = ["None"]
            if not portfolios_df.empty:
                portfolio_opts += portfolios_df['portfolio_name'].tolist()
            selected_portfolio = fc3.selectbox("Filter by Portfolio", portfolio_opts)
            filtered = sales_records_df[sales_records_df['Year'].isin(selected_years)].copy()
            if selected_portfolio != "None" and not portfolios_df.empty:
                pf_row = portfolios_df[portfolios_df['portfolio_name'] == selected_portfolio].iloc[0]
                pf_products = pf_row['products'] if isinstance(pf_row['products'], list) else []
                filtered = filtered[filtered['order_description'].isin(pf_products)]
                st.info(f"📁 Viewing portfolio: **{selected_portfolio}** ({len(pf_products)} products)")
            elif selected_skus:
                filtered = filtered[filtered['order_description'].isin(selected_skus)]
            if filtered.empty:
                st.warning("No data for selected filters.")
            else:
                # --- 1. Revenue Over Time (Monthly, by Year) ---
                st.write("---")
                st.markdown("#### Monthly Revenue Trend")
                monthly = filtered.groupby(['Year', 'MonthNum']).agg(Revenue=('gross_revenue', 'sum')).reset_index()
                monthly['Month'] = monthly['MonthNum'].apply(lambda m: datetime(2000, int(m), 1).strftime('%b'))
                import altair as alt
                rev_chart = alt.Chart(monthly).mark_line(point=True, strokeWidth=3).encode(
                    x=alt.X('Month:N', sort=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'], title='Month'),
                    y=alt.Y('Revenue:Q', title='Revenue ($)'),
                    color=alt.Color('Year:N', title='Year'),
                    tooltip=['Year:N', 'Month:N', alt.Tooltip('Revenue:Q', format='$,.2f')]
                ).properties(height=350)
                st.altair_chart(rev_chart, use_container_width=True)

                # --- 2. Best Sellers (Units Sold) ---
                st.write("---")
                col_bs1, col_bs2 = st.columns(2)
                with col_bs1:
                    st.markdown("#### Best Sellers by Units")
                    units_by_sku = filtered.groupby('order_description').agg(Units=('quantity', 'sum')).sort_values('Units', ascending=False).reset_index()
                    units_chart = alt.Chart(units_by_sku).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                        x=alt.X('Units:Q', title='Units Sold'),
                        y=alt.Y('order_description:N', sort='-x', title=''),
                        tooltip=['order_description:N', 'Units:Q']
                    ).properties(height=max(200, len(units_by_sku) * 35))
                    st.altair_chart(units_chart, use_container_width=True)

                # --- 3. Revenue by SKU ---
                with col_bs2:
                    st.markdown("#### Revenue by Product")
                    rev_by_sku = filtered.groupby('order_description').agg(Revenue=('gross_revenue', 'sum')).sort_values('Revenue', ascending=False).reset_index()
                    rev_sku_chart = alt.Chart(rev_by_sku).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                        x=alt.X('Revenue:Q', title='Revenue ($)'),
                        y=alt.Y('order_description:N', sort='-x', title=''),
                        tooltip=['order_description:N', alt.Tooltip('Revenue:Q', format='$,.2f')]
                    ).properties(height=max(200, len(rev_by_sku) * 35))
                    st.altair_chart(rev_sku_chart, use_container_width=True)

                # --- 4. Gross Margin by SKU ---
                st.write("---")
                st.markdown("#### Gross Margin by Product")
                margin_data = filtered.groupby('order_description').agg(Revenue=('gross_revenue', 'sum'), Profit=('net_profit', 'sum'), COGS=('cogs', 'sum')).reset_index()
                margin_data['Margin %'] = ((margin_data['Profit'] / margin_data['Revenue']) * 100).round(1)
                margin_data = margin_data.sort_values('Margin %', ascending=False)
                margin_chart = alt.Chart(margin_data).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                    x=alt.X('Margin %:Q', title='Gross Margin (%)'),
                    y=alt.Y('order_description:N', sort='-x', title=''),
                    color=alt.condition(alt.datum['Margin %'] >= 50, alt.value('#10b981'), alt.value('#f59e0b')),
                    tooltip=['order_description:N', alt.Tooltip('Margin %:Q', format='.1f'), alt.Tooltip('Revenue:Q', format='$,.2f'), alt.Tooltip('Profit:Q', format='$,.2f')]
                ).properties(height=max(200, len(margin_data) * 35))
                st.altair_chart(margin_chart, use_container_width=True)

                # --- 5. Revenue by Channel ---
                st.write("---")
                col_ch1, col_ch2 = st.columns(2)
                with col_ch1:
                    st.markdown("#### Revenue by Channel")
                    channel_data = filtered.groupby('channel').agg(Revenue=('gross_revenue', 'sum')).sort_values('Revenue', ascending=False).reset_index()
                    channel_chart = alt.Chart(channel_data).mark_arc(innerRadius=50).encode(
                        theta=alt.Theta('Revenue:Q'),
                        color=alt.Color('channel:N', title='Channel'),
                        tooltip=['channel:N', alt.Tooltip('Revenue:Q', format='$,.2f')]
                    ).properties(height=300)
                    st.altair_chart(channel_chart, use_container_width=True)

                # --- 6. Top Clients ---
                with col_ch2:
                    st.markdown("#### Top Clients by Revenue")
                    client_rev = filtered.groupby('account').agg(Revenue=('gross_revenue', 'sum'), Orders=('order_ref_number', 'nunique')).sort_values('Revenue', ascending=False).head(10).reset_index()
                    st.dataframe(client_rev.rename(columns={'account': 'Client'}), use_container_width=True, hide_index=True, column_config={"Revenue": st.column_config.NumberColumn(format="$%.2f")})

                # --- 7. KPI Summary Table ---
                st.write("---")
                st.markdown("#### Product Performance Summary")
                perf = filtered.groupby('order_description').agg(
                    Units=('quantity', 'sum'),
                    Revenue=('gross_revenue', 'sum'),
                    COGS=('cogs', 'sum'),
                    Profit=('net_profit', 'sum'),
                    Orders=('order_ref_number', 'nunique'),
                    Avg_Price=('unit_price', 'mean')
                ).reset_index()
                perf['Margin %'] = ((perf['Profit'] / perf['Revenue']) * 100).round(1)
                perf['Avg Price'] = perf['Avg_Price'].round(2)
                perf = perf.sort_values('Revenue', ascending=False)
                st.dataframe(perf[['order_description', 'Units', 'Revenue', 'COGS', 'Profit', 'Margin %', 'Orders', 'Avg Price']].rename(columns={'order_description': 'Product'}), use_container_width=True, hide_index=True, column_config={
                    "Revenue": st.column_config.NumberColumn(format="$%.2f"),
                    "COGS": st.column_config.NumberColumn(format="$%.2f"),
                    "Profit": st.column_config.NumberColumn(format="$%.2f"),
                    "Avg Price": st.column_config.NumberColumn(format="$%.2f"),
                    "Margin %": st.column_config.NumberColumn(format="%.1f%%")
                })

                # --- 8. Year-over-Year Comparison ---
                if len(selected_years) >= 2:
                    st.write("---")
                    st.markdown("#### Year-over-Year Comparison")
                    yoy = filtered.groupby('Year').agg(Revenue=('gross_revenue', 'sum'), Profit=('net_profit', 'sum'), Units=('quantity', 'sum')).reset_index()
                    yoy_cols = st.columns(len(yoy))
                    for i, (_, row) in enumerate(yoy.iterrows()):
                        with yoy_cols[i]:
                            st.metric(f"{int(row['Year'])} Revenue", f"${row['Revenue']:,.2f}")
                            st.metric(f"{int(row['Year'])} Profit", f"${row['Profit']:,.2f}")
                            st.metric(f"{int(row['Year'])} Units", f"{int(row['Units']):,}")
        else:
            st.info("No sales data available for analytics.")

    # --- 4. RAW MATERIAL LIBRARY ---
    elif menu == "Raw Material Library":
        d = load_tables('inventory')
        inventory = d['inventory']
        st.title("Raw Material Library")
        st.markdown("<p style='opacity: 0.6;'>Manage essential oils, carriers, and active ingredients. Select a material to view its Lot Tracking.</p>", unsafe_allow_html=True)
        if not inventory.empty:
            display_inv = inventory.copy(); display_inv['Cost/g ($)'] = (display_inv['price_per_kg'] / 1000).map('${:,.4f}'.format); display_inv.insert(0, '🔍', False) 
            with st.container(border=True):
                edited_inv = st.data_editor(display_inv[['🔍', 'rm_code', 'trade_name', 'inci_name', 'price_per_kg', 'Cost/g ($)', 'quantity_kg']], use_container_width=True, hide_index=True, disabled=['rm_code', 'Cost/g ($)', 'quantity_kg'])
                if st.button("💾 Synchronize Vault"):
                    for idx, row in edited_inv.iterrows():
                        orig = inventory.loc[idx]
                        if row['trade_name'] != orig['trade_name'] or row['inci_name'] != orig['inci_name'] or row['price_per_kg'] != orig['price_per_kg']:
                            supabase.table('inventory').update({"trade_name": row['trade_name'], "inci_name": row['inci_name'], "price_per_kg": row['price_per_kg']}).eq('id', int(orig['id'])).execute()
                    clear_cache(); st.rerun()
            selected_mats = edited_inv[edited_inv['🔍'] == True]
            if not selected_mats.empty:
                mat = inventory.loc[selected_mats.index[0]]; st.write("##")
                with st.container(border=True):
                    st.markdown(f"#### {mat['trade_name']}")
                    # Calculate Avg and Last cost from lots
                    lots_for_calc = mat.get('lots', [])
                    if isinstance(lots_for_calc, float) or (isinstance(lots_for_calc, str) and lots_for_calc in ["", "nan", "[]"]):
                        lots_for_calc = []
                    avg_cost = float(mat['price_per_kg'])
                    last_cost = float(mat['price_per_kg'])
                    if lots_for_calc:
                        # Backfill price to lots missing it
                        for l in lots_for_calc:
                            if 'Price/Kg' not in l or l.get('Price/Kg') in [None, '', 0]:
                                l['Price/Kg'] = float(mat['price_per_kg'])
                        total_value = sum(float(l.get('Qty (Kg)', 0)) * float(l.get('Price/Kg', mat['price_per_kg'])) for l in lots_for_calc)
                        total_qty = sum(float(l.get('Qty (Kg)', 0)) for l in lots_for_calc)
                        avg_cost = (total_value / total_qty) if total_qty > 0 else float(mat['price_per_kg'])
                        # Last cost = most recent lot by Rcv Date
                        sorted_lots = sorted(lots_for_calc, key=lambda x: str(x.get('Rcv Date', '')), reverse=True)
                        if sorted_lots:
                            last_cost = float(sorted_lots[0].get('Price/Kg', mat['price_per_kg']))
                    cost_delta_pct = ((last_cost - avg_cost) / avg_cost * 100) if avg_cost > 0 else 0
                    c1, c2, c3, c4 = st.columns(4)
                    c1.write(f"**Code:** {mat['rm_code']}<br>**INCI:** {mat['inci_name']}", unsafe_allow_html=True)
                    c2.write(f"**Total Stock:** {mat['quantity_kg']} Kg<br>**Shelf Value:** ${(avg_cost * float(mat['quantity_kg'])):.2f}", unsafe_allow_html=True)
                    c3.metric("Avg Cost/Kg", f"${avg_cost:.2f}")
                    c4.metric("Last Cost/Kg", f"${last_cost:.2f}", f"{cost_delta_pct:+.1f}% vs avg" if abs(cost_delta_pct) > 0.1 else None)
                    if abs(cost_delta_pct) > 10:
                        st.warning(f"⚠️ Last purchase was {cost_delta_pct:+.1f}% vs average. Consider updating retail pricing or renegotiating with supplier.")
                    st.write("---")
                    st.markdown("#### 📦 Lot Tracking Ledgers")
                    lots = mat.get('lots', [])
                    if isinstance(lots, float): lots = []
                    elif isinstance(lots, str) and lots in ["", "nan", "[]"]: lots = []
                    if not lots:
                        today_str = datetime.today().strftime('%Y-%m-%d')
                        exp_str = (datetime.today() + pd.DateOffset(years=2)).strftime('%Y-%m-%d')
                        lots = [{"Lot Number": f"{mat['rm_code']}-L01", "Mfg Date": today_str, "Rcv Date": today_str, "Exp Date": exp_str, "Qty (Kg)": float(mat['quantity_kg']), "Price/Kg": float(mat['price_per_kg']), "Current": True}]
                    # Backfill price field on existing lots
                    for l in lots:
                        if 'Price/Kg' not in l:
                            l['Price/Kg'] = float(mat['price_per_kg'])
                    lots_df = pd.DataFrame(lots)
                    with st.form(f"lots_form_{mat['id']}"):
                        st.info("💡 Each lot tracks its own purchase price. Average and Last costs are auto-calculated above.")
                        ed_lots = st.data_editor(lots_df, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Current": st.column_config.CheckboxColumn("Current Lot", default=False), "Mfg Date": st.column_config.TextColumn("Mfg Date (YYYY-MM-DD)"), "Rcv Date": st.column_config.TextColumn("Rcv Date (YYYY-MM-DD)"), "Exp Date": st.column_config.TextColumn("Exp Date (YYYY-MM-DD)"), "Qty (Kg)": st.column_config.NumberColumn("Qty (Kg)", format="%.3f"), "Price/Kg": st.column_config.NumberColumn("Price/Kg ($)", format="$%.2f", min_value=0.0)})
                        if st.form_submit_button("💾 Save Lots & Update Total Stock", type="primary"):
                            current_count = ed_lots['Current'].sum() if 'Current' in ed_lots.columns else 0
                            if current_count > 1: st.error("⚠️ Only one lot can be marked as the 'Current' lot.")
                            else:
                                new_lots_json = []
                                for _, lrow in ed_lots.iterrows():
                                    clean_lot = {}
                                    for k, v in lrow.items():
                                        if pd.isna(v): clean_lot[k] = None
                                        elif isinstance(v, bool): clean_lot[k] = bool(v)
                                        elif hasattr(v, 'item'): clean_lot[k] = v.item()
                                        else: clean_lot[k] = v
                                    new_lots_json.append(clean_lot)
                                new_total_kg = float(ed_lots['Qty (Kg)'].sum()) if 'Qty (Kg)' in ed_lots.columns else 0.0
                                new_avg = 0.0
                                new_last = float(mat['price_per_kg'])
                                if new_total_kg > 0:
                                    total_val = sum(float(l.get('Qty (Kg)', 0) or 0) * float(l.get('Price/Kg', 0) or 0) for l in new_lots_json)
                                    new_avg = total_val / new_total_kg
                                    sorted_new = sorted(new_lots_json, key=lambda x: str(x.get('Rcv Date', '')), reverse=True)
                                    if sorted_new: new_last = float(sorted_new[0].get('Price/Kg', mat['price_per_kg']) or mat['price_per_kg'])
                                supabase.table('inventory').update({"lots": new_lots_json, "quantity_kg": new_total_kg, "price_per_kg": new_last}).eq('id', int(mat['id'])).execute()
                                st.success(f"Saved! Avg: ${new_avg:.2f}/Kg | Last: ${new_last:.2f}/Kg")
                                if "cogs_synced_this_session" in st.session_state:
                                    del st.session_state.cogs_synced_this_session
                                time.sleep(1.5); clear_cache(); st.rerun()
                    with st.expander("System Actions"):
                        del_pass = st.text_input("Authorization Passcode", type="password", key="dmp")
                        if st.button("Erase Record") and del_pass == "lab2026":
                            supabase.table('inventory').delete().eq('id', int(mat['id'])).execute(); clear_cache(); st.rerun()
        st.write("---")
        st.info("💡 Need to add multiple materials at once? Go to **🛠️ Admin Tools → Bulk Import** to upload a CSV.")
        with st.expander("➕ Register New Material"):
            with st.form("add_material", clear_on_submit=True):
                c1, c2 = st.columns(2); new_t = c1.text_input("Trade Name"); new_i = c1.text_input("INCI Name"); new_p = c2.number_input("Price/Kg ($)", min_value=0.0); new_q = c2.number_input("Initial Qty (Kg)", min_value=0.0)
                st.markdown("**Initial Lot Details**")
                l1, l2, l3, l4 = st.columns(4)
                new_lot = l1.text_input("Lot Number", "L-01")
                new_mfg = l2.date_input("Mfg Date")
                new_rcv = l3.date_input("Rcv Date")
                new_exp = l4.date_input("Exp Date", value=datetime.today() + pd.DateOffset(years=2))
                if st.form_submit_button("Register") and new_t != "":
                    next_id = 1 if inventory.empty else int(inventory['id'].max()) + 1
                    rm_code = f"RM{next_id:05d}"
                    init_lot = [{"Lot Number": new_lot, "Mfg Date": new_mfg.strftime('%Y-%m-%d'), "Rcv Date": new_rcv.strftime('%Y-%m-%d'), "Exp Date": new_exp.strftime('%Y-%m-%d'), "Qty (Kg)": float(new_q), "Current": True}]
                    supabase.table('inventory').insert({"rm_code": rm_code, "trade_name": new_t, "inci_name": new_i, "price_per_kg": new_p, "quantity_kg": new_q, "lots": init_lot}).execute(); clear_cache(); st.rerun()

    # --- 5. PACKAGING LIBRARY ---
    elif menu == "Packaging Library":
        d = load_tables('packaging')
        packaging = d['packaging']
        st.title("Packaging Library")
        st.markdown("<p style='opacity: 0.6;'>Track bottles, droppers, caps, and labels. Select a material to view its Lot Tracking.</p>", unsafe_allow_html=True)
        if not packaging.empty:
            display_pk = packaging.copy(); display_pk.insert(0, '🔍', False)
            with st.container(border=True):
                edited_pk = st.data_editor(display_pk[['🔍', 'pm_code', 'material_name', 'supplier', 'cost_per_unit', 'remaining_quantity']], use_container_width=True, hide_index=True, disabled=['pm_code', 'remaining_quantity'])
                if st.button("💾 Synchronize Vault"):
                    for idx, row in edited_pk.iterrows():
                        orig = packaging.loc[idx]
                        if row['material_name'] != orig['material_name'] or row['supplier'] != orig['supplier'] or row['cost_per_unit'] != orig['cost_per_unit']:
                            supabase.table('packaging').update({"material_name": row['material_name'], "supplier": row['supplier'], "cost_per_unit": row['cost_per_unit']}).eq('id', int(orig['id'])).execute()
                    clear_cache(); st.rerun()
            selected_pk = edited_pk[edited_pk['🔍'] == True]
            if not selected_pk.empty:
                p_mat = packaging.loc[selected_pk.index[0]]; st.write("##")
                with st.container(border=True):
                    st.markdown(f"#### {p_mat['material_name']}")
                    st.write(f"**Code:** {p_mat['pm_code']} | **Supplier:** {p_mat['supplier']} | **Total Stock:** {p_mat['remaining_quantity']} Units")
                    st.write("---")
                    st.markdown("#### 📦 Lot Tracking Ledgers")
                    lots = p_mat.get('lots', [])
                    if isinstance(lots, float): lots = []
                    elif isinstance(lots, str) and lots in ["", "nan", "[]"]: lots = []
                    if not lots:
                        today_str = datetime.today().strftime('%Y-%m-%d')
                        lots = [{"Lot Number": f"{p_mat['pm_code']}-L01", "Rcv Date": today_str, "Qty (Units)": int(p_mat['remaining_quantity']), "Current": True}]
                    lots_df = pd.DataFrame(lots)
                    with st.form(f"pk_lots_form_{p_mat['id']}"):
                        st.info("💡 Edit quantities, add new lots, and mark exactly ONE lot as 'Current Lot'. Total Stock will auto-update.")
                        ed_lots = st.data_editor(lots_df, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Current": st.column_config.CheckboxColumn("Current Lot", default=False), "Rcv Date": st.column_config.TextColumn("Rcv Date (YYYY-MM-DD)"), "Qty (Units)": st.column_config.NumberColumn("Qty (Units)", step=1)})
                        if st.form_submit_button("💾 Save Lots & Update Total Stock", type="primary"):
                            current_count = ed_lots['Current'].sum() if 'Current' in ed_lots.columns else 0
                            if current_count > 1: st.error("⚠️ Only one lot can be marked as the 'Current' lot.")
                            else:
                                new_lots_json = []
                                for _, lrow in ed_lots.iterrows():
                                    clean_lot = {}
                                    for k, v in lrow.items():
                                        if pd.isna(v): clean_lot[k] = None
                                        elif isinstance(v, bool): clean_lot[k] = bool(v)
                                        elif hasattr(v, 'item'): clean_lot[k] = v.item()
                                        else: clean_lot[k] = v
                                    new_lots_json.append(clean_lot)
                                new_total_qty = int(ed_lots['Qty (Units)'].sum()) if 'Qty (Units)' in ed_lots.columns else 0
                                supabase.table('packaging').update({"lots": new_lots_json, "remaining_quantity": new_total_qty}).eq('id', int(p_mat['id'])).execute()
                                st.success("Lots updated successfully! Total Stock recalculated.")
                                time.sleep(1.5); clear_cache(); st.rerun()
                    with st.expander("System Actions"):
                        if st.button("Erase Record") and st.text_input("Authorization", type="password", key="dpp") == "lab2026":
                            supabase.table('packaging').delete().eq('id', int(p_mat['id'])).execute(); clear_cache(); st.rerun()
        st.write("---")
        st.info("💡 Need to add multiple packaging items at once? Go to **🛠️ Admin Tools → Bulk Import** to upload a CSV.")
        with st.expander("➕ Register New Packaging"):
            with st.form("add_packaging", clear_on_submit=True):
                c1, c2 = st.columns(2); p_n = c1.text_input("Material Name"); p_s = c1.text_input("Supplier"); p_c = c2.number_input("Cost/Unit ($)", min_value=0.0); p_q = c2.number_input("Initial Qty", min_value=0, step=1)
                st.markdown("**Initial Lot Details**")
                l1, l2 = st.columns(2)
                new_lot = l1.text_input("Lot Number", "L-01")
                new_rcv = l2.date_input("Rcv Date")
                if st.form_submit_button("Register") and p_n != "":
                    next_pm = 1 if packaging.empty else int(packaging['id'].max()) + 1
                    pm_code = f"PM{next_pm:05d}"
                    init_lot = [{"Lot Number": new_lot, "Rcv Date": new_rcv.strftime('%Y-%m-%d'), "Qty (Units)": int(p_q), "Current": True}]
                    supabase.table('packaging').insert({"pm_code": pm_code, "material_name": p_n, "supplier": p_s, "cost_per_unit": p_c, "remaining_quantity": p_q, "lots": init_lot}).execute(); clear_cache(); st.rerun()

    # --- 6. FINISHED PRODUCTS LIBRARY ---
    elif menu == "Finished Products":
        d = load_tables('finished_goods', 'cogs_records')
        finished_goods = d['finished_goods']; cogs_records_df = d['cogs_records']
        st.title("Finished Products")
        st.markdown("<p style='opacity: 0.6;'>Manage retail-ready inventory directly from your saved COGS profiles.</p>", unsafe_allow_html=True)
        if not finished_goods.empty:
            display_fp = finished_goods.copy()
            display_fp.insert(0, '🔍', False)
            st.write("💡 *Edit stock quantities directly in the table below.*")
            with st.container(border=True):
                edited_fp = st.data_editor(display_fp[['🔍', 'fp_code', 'product_name', 'stock_quantity', 'unit_cogs', 'retail_price']], use_container_width=True, hide_index=True, disabled=['fp_code', 'unit_cogs', 'retail_price'], column_config={"unit_cogs": st.column_config.NumberColumn("Unit COGS", format="$%.2f"), "retail_price": st.column_config.NumberColumn("Retail Price", format="$%.2f")})
                if st.button("💾 Synchronize Vault", type="primary"):
                    for idx, row in edited_fp.iterrows():
                        orig = finished_goods.loc[idx]
                        if row['stock_quantity'] != orig['stock_quantity']:
                            supabase.table('finished_products').update({"stock_quantity": row['stock_quantity']}).eq('id', int(orig['id'])).execute()
                    st.success("Finished goods synced!")
                    clear_cache(); st.rerun()
            selected_fp = edited_fp[edited_fp['🔍'] == True]
            if not selected_fp.empty:
                fp_item = finished_goods.loc[selected_fp.index[0]]
                st.write("##")
                with st.container(border=True):
                    st.markdown(f"#### {fp_item['product_name']}")
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Code:** {fp_item['fp_code']}")
                    c2.write(f"**In Stock:** {fp_item['stock_quantity']} Units")
                    margin = ((fp_item['retail_price'] - fp_item['unit_cogs']) / fp_item['retail_price'] * 100) if fp_item['retail_price'] > 0 else 0
                    c3.write(f"**Profit Margin:** {margin:.1f}%")
                    with st.expander("System Actions"):
                        if st.button("Erase Record") and st.text_input("Authorization Passcode", type="password", key="dfpp") == "lab2026":
                            supabase.table('finished_products').delete().eq('id', int(fp_item['id'])).execute(); clear_cache(); st.rerun()
        else:
            st.info("No finished products currently in stock.")
        st.write("---")
        with st.expander("➕ Log New Finished Product Batch"):
            if not cogs_records_df.empty:
                with st.form("add_fp", clear_on_submit=True):
                    c1, c2 = st.columns([2, 1])
                    cogs_opts = [f"[{r['id']}] {r['product_name']}" for _, r in cogs_records_df.iterrows()]
                    sel_cogs = c1.selectbox("Select Target Product (From COGS Vault)", cogs_opts)
                    fp_q = c2.number_input("Bottles Produced (Qty)", min_value=1, value=10, step=1)
                    if st.form_submit_button("Add to Stock"):
                        cogs_id = int(sel_cogs.split("]")[0].replace("[", ""))
                        matched_cogs = cogs_records_df[cogs_records_df['id'] == cogs_id].iloc[0]
                        target_name = matched_cogs['product_name']
                        target_cogs = float(matched_cogs['total_cogs'])
                        target_retail = float(matched_cogs['target_retail'])
                        if not finished_goods.empty and target_name in finished_goods['product_name'].values:
                            existing_product = finished_goods[finished_goods['product_name'] == target_name]
                            existing_id = int(existing_product.iloc[0]['id'])
                            new_qty = int(existing_product.iloc[0]['stock_quantity']) + fp_q
                            supabase.table('finished_products').update({"stock_quantity": new_qty, "unit_cogs": target_cogs, "retail_price": target_retail}).eq('id', existing_id).execute()
                        else:
                            next_fp = 1 if finished_goods.empty else int(finished_goods['id'].max()) + 1
                            supabase.table('finished_products').insert({"fp_code": f"FP{next_fp:05d}", "product_name": target_name, "stock_quantity": fp_q, "unit_cogs": target_cogs, "retail_price": target_retail}).execute()
                        clear_cache(); st.rerun()
            else:
                st.warning("⚠️ You need to architect and save a product profile in the **COGS Calculator** before you can log it to your finished inventory.")

    # --- PURCHASE REQUISITION ---
    elif menu == "Purchase Requisition":
        d = load_tables('inventory')
        inventory = d['inventory']
        st.title("Purchase Requisition")
        st.markdown("<p style='opacity: 0.6;'>Tick items to build your procurement list.</p>", unsafe_allow_html=True)

        if not inventory.empty:
            rm_df = inventory[['rm_code', 'trade_name', 'quantity_kg']].copy()
            rm_df.insert(0, 'Select', False)
            rm_df['Order Qty (Kg)'] = 1.0

            result = st.data_editor(
                rm_df,
                use_container_width=True,
                hide_index=True,
                disabled=['rm_code', 'trade_name', 'quantity_kg'],
                key=f"pr_table_{st.session_state.get('pr_table_version', 0)}",
                column_config={
                    "rm_code": "Code",
                    "trade_name": "Material",
                    "quantity_kg": "Current Stock (Kg)",
                    "Order Qty (Kg)": st.column_config.NumberColumn("Order Qty (Kg)", min_value=0.1, step=0.5)
                }
            )

            selected = result[result['Select'] == True]

            st.write("---")
            if not selected.empty:
                st.markdown(f"#### 📋 Selected Items ({len(selected)})")
                st.dataframe(selected[['trade_name', 'quantity_kg', 'Order Qty (Kg)']].rename(columns={'trade_name': 'Material', 'quantity_kg': 'Current Stock', 'Order Qty (Kg)': 'Order Qty'}), use_container_width=True, hide_index=True)

                if st.button("✉️ Generate Supplier Message", type="primary"):
                    items_text = ""
                    for _, row in selected.iterrows():
                        items_text += f"- {row['trade_name']} ({row['Order Qty (Kg)']:.1f} Kg)\n"
                    msg = f"Hi, I kindly need the below items:\n\n{items_text}\nPlease confirm availability and lead time. Thank you."
                    st.session_state.pr_generated_msg = msg

                if "pr_generated_msg" in st.session_state:
                    st.write("---")
                    editable_message = st.text_area("Edit Message", value=st.session_state.pr_generated_msg, height=200, key="pr_editable")
                    mc1, mc2 = st.columns(2)
                    if mc1.button("📋 Show Copyable Text", use_container_width=True):
                        st.code(editable_message, language=None)
                    wa_url = f"https://wa.me/?text={editable_message.replace(chr(10), '%0A').replace(' ', '%20').replace('#', '%23')}"
                    mc2.link_button("💬 Send via WhatsApp", wa_url, use_container_width=True)
                st.write("---")
                if st.button("🗑️ Clear All Selections"):
                    st.session_state.pr_table_version = st.session_state.get("pr_table_version", 0) + 1
                    for key in list(st.session_state.keys()):
                        if key.startswith("pr_") and key != "pr_table_version":
                            del st.session_state[key]
                    st.rerun()
            else:
                st.info("Tick the checkbox next to items you want to order.")
                if "pr_generated_msg" in st.session_state:
                    del st.session_state.pr_generated_msg

    # --- 7. FORMULA LIBRARY ---
    elif menu == "Formula Library":
        d = load_tables('formulas', 'inventory')
        formulas_df = d['formulas']; inventory = d['inventory']
        st.title("📚 Formula Library")
        st.markdown("<p style='opacity: 0.6;'>Inspect read-only recipes and execute live manufacturing batches.</p>", unsafe_allow_html=True)
        if not formulas_df.empty:
            formulas_df['base_code'] = formulas_df['fr_code'].apply(lambda x: str(x).split('-')[0])
            summary_df = formulas_df.sort_values(by='fr_code').drop_duplicates(subset=['base_code'], keep='first').copy()
            summary_df['Family Name'] = summary_df['formula_name'].apply(lambda x: re.sub(r' V\d+$', '', str(x)))
            st.write("💡 *Select a Formula Family below to inspect its editions and execute production.*")
            f_select = st.dataframe(summary_df[['base_code', 'Family Name']], use_container_width=True, hide_index=True, on_select="rerun", selection_mode="single-row")
            if f_select.selection.rows:
                sel_base = summary_df.iloc[f_select.selection.rows[0]]['base_code']
                family_editions = formulas_df[formulas_df['base_code'] == sel_base].sort_values(by='fr_code', ascending=False)
                st.write("##")
                with st.container(border=True):
                    if len(family_editions) > 1:
                        edition_opts = [f"[{r['fr_code']}] {r['formula_name']}" for _, r in family_editions.iterrows()]
                        sel_edition_str = st.selectbox("📌 Select Specific Edition:", edition_opts)
                        sel_fr_code = sel_edition_str.split("]")[0].replace("[", "")
                        sel_f = family_editions[family_editions['fr_code'] == sel_fr_code].iloc[0]
                    else:
                        sel_f = family_editions.iloc[0]
                        st.markdown(f"#### ⚗️ {sel_f['fr_code']} - {sel_f['formula_name']}")
                    recipe_data = sel_f['recipe']
                    if isinstance(recipe_data, dict):
                        recipe_items = [{"Phase": "A", "Ingredient": k, "%": v} for k, v in recipe_data.items()]
                    elif isinstance(recipe_data, list):
                        recipe_items = []
                        for item in recipe_data:
                            if "Phase" not in item: item["Phase"] = "A"
                            recipe_items.append(item)
                    else:
                        recipe_items = []
                    st.write("---")
                    b_size = st.number_input("Target Batch Size (grams)", min_value=1.0, value=1000.0, step=100.0)
                    st.write("---")
                    calc_data = []; stock_ok = True; total_cost = 0.0
                    for row in recipe_items:
                        ing = row.get('Ingredient'); p = row.get('%', 0); phase = row.get('Phase', 'A')
                        req_g = (p/100) * b_size
                        m = inventory[inventory['trade_name'] == ing]
                        if not m.empty:
                            s_kg = float(m['quantity_kg'].values[0]); p_kg = float(m['price_per_kg'].values[0])
                            rm_c = str(m['rm_code'].values[0])
                            if s_kg < (req_g/1000): stock_ok = False
                            cost = (req_g/1000)*p_kg; total_cost += cost
                            calc_data.append({"Phase": phase, "RM Code": rm_c, "Material": ing, "Formula %": f"{p}%", "Needed (g)": f"{req_g:.2f}", "Stock Status": "✅ Available" if s_kg >= (req_g/1000) else "❌ Shortage", "Est. Cost": f"${cost:.4f}", "req_kg": req_g/1000, "stock_kg": s_kg})
                        else:
                            stock_ok = False
                            calc_data.append({"Phase": phase, "RM Code": "N/A", "Material": ing, "Formula %": f"{p}%", "Needed (g)": f"{req_g:.2f}", "Stock Status": "⚠️ Not in Vault", "Est. Cost": "$0.00", "req_kg": 0, "stock_kg": 0})
                    calc_df = pd.DataFrame(calc_data)
                    if not calc_df.empty:
                        st.dataframe(calc_df.sort_values(by="Phase")[['Phase', 'RM Code', 'Material', 'Formula %', 'Needed (g)', 'Stock Status', 'Est. Cost']], use_container_width=True, hide_index=True)
                    st.write("---")
                    st.markdown("#### 📋 Manufacturing Procedure")
                    proc_text = sel_f.get('procedure', 'No written procedure documented for this formula.')
                    if pd.isna(proc_text) or str(proc_text).strip() == "": proc_text = "No written procedure documented for this formula."
                    st.info(proc_text)
                    st.write("---")
                    col_cost, col_btn = st.columns([1, 1])
                    col_cost.metric("Projected Batch Cost", f"${total_cost:.2f}")
                    with col_btn:
                        st.write("<br>", unsafe_allow_html=True)
                        if st.button("🚀 Execute Production", type="primary", use_container_width=True):
                            if stock_ok:
                                l_r = supabase.table('production_records').select("id").order("id", desc=True).limit(1).execute()
                                n_id = 1 if not l_r.data else l_r.data[0]['id'] + 1
                                b_no, l_no = f"B-{n_id:05d}", f"LOT-{datetime.now().strftime('%Y%m%d')}-{n_id:02d}"
                                for d in calc_data:
                                    supabase.table('inventory').update({'quantity_kg': d['stock_kg'] - d['req_kg']}).eq('trade_name', d['Material']).execute()
                                supabase.table('production_records').insert({"fr_code": sel_f['fr_code'], "formula_name": sel_f['formula_name'], "batch_number": b_no, "lot_number": l_no, "batch_size_g": b_size, "total_cost": total_cost}).execute()
                                st.balloons(); clear_cache(); st.rerun()
                            else: st.error("Cannot produce: Material Shortage detected.")
                    st.divider()
                    if st.session_state.get("user_role") != "admin":
                        st.info("🔒 Read-only mode. Contact admin to edit formulas.")
                    else:
                        c_act1, c_act2, c_act3, c_act4 = st.columns(4)
                        with c_act1:
                            with st.expander("✏️ Edit Current"):
                                if st.button("Edit Edition", use_container_width=True):
                                    st.session_state.builder = pd.DataFrame(recipe_items)
                                    st.session_state.draft_name = sel_f['formula_name']
                                    st.session_state.draft_procedure = str(proc_text) if proc_text != "No written procedure documented for this formula." else ""
                                    st.session_state.edit_formula_id = int(sel_f['id'])
                                    st.session_state.edit_fr_code = sel_f['fr_code']
                                    if "base_fr_code" in st.session_state: del st.session_state["base_fr_code"]
                                    st.session_state.active_nav = "Formula Builder"
                                    clear_cache(); st.rerun()
                        with c_act2:
                            with st.expander("🔄 New Edition"):
                                if st.button("Draft Version", use_container_width=True):
                                    st.session_state.builder = pd.DataFrame(recipe_items)
                                    match = re.search(r' V(\d+)$', sel_f['formula_name'])
                                    if match:
                                        new_v = int(match.group(1)) + 1
                                        new_name = re.sub(r' V\d+$', f' V{new_v}', sel_f['formula_name'])
                                    else:
                                        new_name = f"{sel_f['formula_name']} V2"
                                    st.session_state.draft_name = new_name
                                    st.session_state.base_fr_code = sel_f['fr_code']
                                    st.session_state.draft_procedure = str(proc_text) if proc_text != "No written procedure documented for this formula." else ""
                                    if "edit_formula_id" in st.session_state: del st.session_state["edit_formula_id"]
                                    st.session_state.active_nav = "Formula Builder"
                                    clear_cache(); st.rerun()
                        with c_act3:
                            with st.expander("📋 Duplicate"):
                                if st.button("Copy to New Base", use_container_width=True):
                                    st.session_state.builder = pd.DataFrame(recipe_items)
                                    st.session_state.draft_name = f"{sel_f['formula_name']} (Copy)"
                                    st.session_state.draft_procedure = str(proc_text) if proc_text != "No written procedure documented for this formula." else ""
                                    for k in ["edit_formula_id", "edit_fr_code", "base_fr_code"]:
                                        if k in st.session_state: del st.session_state[k]
                                    st.session_state.active_nav = "Formula Builder"
                                    clear_cache(); st.rerun()
                        with c_act4:
                            with st.expander("🗑️ Erase"):
                                del_f_pass = st.text_input("Authorization Passcode", type="password", key="dfp")
                                if st.button("Permanently Delete", use_container_width=True) and del_f_pass == "lab2026":
                                    supabase.table('formulas').delete().eq('id', int(sel_f['id'])).execute(); clear_cache(); st.rerun()
        else:
            st.info("No formulas architected yet.")

    # --- 7.5 FORMULA BUILDER ---
    elif menu == "Formula Builder":
        d = load_tables('formulas', 'inventory')
        formulas_df = d['formulas']; inventory = d['inventory']
        st.title("⚙️ Formula Builder")
        st.markdown("<p style='opacity: 0.6;'>Draft, calculate, and version control your recipes here.</p>", unsafe_allow_html=True)
        c_build, c_metrics = st.columns([3, 2])
        with c_build:
            if "edit_formula_id" in st.session_state:
                st.markdown(f"<span style='opacity: 0.9; font-size: 0.85rem; font-weight: 600;'>✏️ EDITING MODE: Overwriting {st.session_state.edit_fr_code}</span>", unsafe_allow_html=True)
                if st.button("❌ Cancel Edit & Start Fresh"):
                    st.session_state.builder = pd.DataFrame([{"Phase": "A", "Ingredient": None, "%": 0.0}])
                    for key in ["draft_name", "edit_formula_id", "edit_fr_code", "draft_procedure"]:
                        if key in st.session_state: del st.session_state[key]
                    clear_cache(); st.rerun()
                st.write("")
            elif "base_fr_code" in st.session_state:
                base_disp = st.session_state.base_fr_code.split('-')[0]
                st.markdown(f"<span style='opacity: 0.9; font-size: 0.85rem; font-weight: 600;'>🔗 NEW EDITION MODE: Linked to Parent {base_disp}</span>", unsafe_allow_html=True)
                if st.button("❌ Cancel Edition & Start Fresh"):
                    st.session_state.builder = pd.DataFrame([{"Phase": "A", "Ingredient": None, "%": 0.0}])
                    for key in ["draft_name", "base_fr_code", "draft_procedure"]:
                        if key in st.session_state: del st.session_state[key]
                    clear_cache(); st.rerun()
                st.write("")
            f_name = st.text_input("Formula Moniker", value=st.session_state.get("draft_name", ""), placeholder="e.g., Actiflam Hair Growth Oil")
            if "builder" not in st.session_state: 
                st.session_state.builder = pd.DataFrame([{"Phase": "A", "Ingredient": None, "%": 0.0}])
            ing_options = inventory['trade_name'].tolist() if not inventory.empty else ["No materials registered"]
            edit_df = st.data_editor(st.session_state.builder, num_rows="dynamic", use_container_width=True, column_config={"Phase": st.column_config.SelectboxColumn("Phase", options=["A", "B", "C", "D", "E", "F"], required=True), "Ingredient": st.column_config.SelectboxColumn("Ingredient", options=ing_options, required=True)})
            procedure_text = st.text_area("Manufacturing Procedure", value=st.session_state.get("draft_procedure", ""), placeholder="1. Heat Phase A to 75°C...", height=150)
        with c_metrics:
            st.write("<div style='margin-top: 2.2rem;'></div>", unsafe_allow_html=True)
            total_cost_kg = 0.0; live_data = []
            for _, row in edit_df.iterrows():
                ing = row.get('Ingredient'); perc = row.get('%', 0.0); phase = row.get('Phase', 'A')
                if ing and pd.notna(ing) and ing != "None" and not inventory.empty and ing in inventory['trade_name'].values:
                    m_row = inventory[inventory['trade_name'] == ing].iloc[0]
                    price = float(m_row['price_per_kg'])
                    rm_c = str(m_row['rm_code'])
                    cost_contrib = (perc / 100.0) * price; total_cost_kg += cost_contrib
                    live_data.append({"Phase": phase, "RM Code": rm_c, "Material": ing, "Cost": f"${cost_contrib:,.2f}"})
            if live_data: 
                st.dataframe(pd.DataFrame(live_data).sort_values('Phase')[['Phase', 'RM Code', 'Material', 'Cost']], use_container_width=True, hide_index=True)
            else: 
                st.info("Select ingredients to see live costs.")
            st.metric("Total Formula Cost / Kg", f"${total_cost_kg:,.2f}")
            total_perc = edit_df["%"].sum() if "%" in edit_df.columns else 0.0
            if round(total_perc, 2) == 100.0:
                st.success("✅ Formula is balanced (100%)")
                btn_label = "💾 Update Existing Edition" if "edit_formula_id" in st.session_state else "Commit Formula to Vault"
                if st.button(btn_label, type="primary", use_container_width=True) and f_name:
                    recipe_json = edit_df.to_dict(orient='records')
                    if "edit_formula_id" in st.session_state:
                        supabase.table("formulas").update({"formula_name": f_name, "recipe": recipe_json, "procedure": procedure_text}).eq('id', st.session_state.edit_formula_id).execute()
                        st.success("Updated Successfully!")
                    else:
                        if "base_fr_code" in st.session_state:
                            base_code = st.session_state.base_fr_code.split('-')[0]
                            existing_v_count = formulas_df[formulas_df['fr_code'].str.startswith(base_code)].shape[0]
                            fr_c = f"{base_code}-{existing_v_count + 1}"
                        else:
                            if formulas_df.empty:
                                fr_c = "FR00001"
                            else:
                                root_codes = formulas_df['fr_code'].str.extract(r'FR(\d{5})')[0].dropna().astype(int)
                                next_id = root_codes.max() + 1 if not root_codes.empty else 1
                                fr_c = f"FR{next_id:05d}"
                        supabase.table("formulas").insert({"fr_code": fr_c, "formula_name": f_name, "recipe": recipe_json, "procedure": procedure_text}).execute()
                        st.success("Saved to Library!")
                    st.session_state.builder = pd.DataFrame([{"Phase": "A", "Ingredient": None, "%": 0.0}])
                    for key in ["draft_name", "base_fr_code", "draft_procedure", "edit_formula_id", "edit_fr_code"]:
                        if key in st.session_state: del st.session_state[key]
            else: st.warning(f"⚠️ Total: {total_perc}% (Must equal 100%)")

    # --- 8. COGS CALCULATOR ---
    elif menu == "COGS Calculator":
        d = load_tables('formulas', 'inventory', 'packaging', 'cogs_records')
        formulas_df = d['formulas']; inventory = d['inventory']; packaging = d['packaging']; cogs_records_df = d['cogs_records']
        st.title("Cost of Goods Sold")
        st.markdown("<p style='opacity: 0.6;'>Build profiles, monitor margins, and react to RM price changes.</p>", unsafe_allow_html=True)

        # Helper: compute avg and last cost per Kg for a raw material
        def get_rm_costs(mat_row):
            lots = mat_row.get('lots', [])
            if isinstance(lots, float) or (isinstance(lots, str) and lots in ["", "nan", "[]"]):
                lots = []
            default_p = float(mat_row['price_per_kg'])
            if not lots:
                return default_p, default_p
            total_val = sum(float(l.get('Qty (Kg)', 0)) * float(l.get('Price/Kg', default_p)) for l in lots)
            total_q = sum(float(l.get('Qty (Kg)', 0)) for l in lots)
            avg = (total_val / total_q) if total_q > 0 else default_p
            sorted_l = sorted(lots, key=lambda x: str(x.get('Rcv Date', '')), reverse=True)
            last = float(sorted_l[0].get('Price/Kg', default_p)) if sorted_l else default_p
            return avg, last

        # Helper: compute fresh COGS for a saved profile using current RM prices
        def recalc_profile_cogs(cogs_item, basis="avg"):
            recalc_formula_name = cogs_item['formula_name']
            recalc_fill_wt = float(cogs_item['fill_weight_g'])
            recalc_bulk = 0.0
            if not formulas_df.empty and recalc_formula_name in formulas_df['formula_name'].values:
                rec_f = formulas_df[formulas_df['formula_name'] == recalc_formula_name].iloc[0]['recipe']
                if isinstance(rec_f, dict):
                    rec_items_r = [{"Ingredient": k, "%": v} for k, v in rec_f.items()]
                elif isinstance(rec_f, list):
                    rec_items_r = rec_f
                else:
                    rec_items_r = []
                for rr in rec_items_r:
                    r_ing = rr.get('Ingredient'); r_pct = rr.get('%', 0)
                    r_req = (r_pct / 100) * recalc_fill_wt
                    r_m = inventory[inventory['trade_name'] == r_ing]
                    if not r_m.empty:
                        avg_p, last_p = get_rm_costs(r_m.iloc[0])
                        chosen = avg_p if basis == "avg" else last_p
                        recalc_bulk += (r_req / 1000) * chosen
            recalc_pack = float(cogs_item.get('packaging_cost', 0) or 0)
            recalc_mfg = float(cogs_item.get('mfg_cost', 0) or 0)
            recalc_lbl = float(cogs_item.get('label_cost', 0) or 0)
            return recalc_bulk + recalc_pack + recalc_mfg + recalc_lbl, recalc_bulk

        tab_dash, tab_build, tab_profiles = st.tabs(["📊 Dashboard", "🧪 Build New COGS", "📂 Saved Profiles"])

        # ============= TAB 1: DASHBOARD =============
        with tab_dash:
            if not cogs_records_df.empty:
                if 'is_active' in cogs_records_df.columns:
                    active_cogs = cogs_records_df[cogs_records_df['is_active'] != False].copy()
                else:
                    active_cogs = cogs_records_df.copy()

                # Build Avg vs Last comparison for each profile
                dash_data = []
                for _, prof in active_cogs.iterrows():
                    cogs_avg_now, _ = recalc_profile_cogs(prof, "avg")
                    cogs_last_now, _ = recalc_profile_cogs(prof, "last")
                    saved_cogs = float(prof['total_cogs'])
                    retail = float(prof['target_retail'])
                    margin_avg = ((retail - cogs_avg_now) / retail * 100) if retail > 0 else 0
                    margin_last = ((retail - cogs_last_now) / retail * 100) if retail > 0 else 0
                    margin_saved = ((retail - saved_cogs) / retail * 100) if retail > 0 else 0
                    drift = cogs_avg_now - saved_cogs
                    dash_data.append({
                        "Product": prof['product_name'],
                        "Retail": retail,
                        "Saved COGS": saved_cogs,
                        "COGS @ Avg (Now)": cogs_avg_now,
                        "COGS @ Last (Now)": cogs_last_now,
                        "Margin @ Avg": margin_avg,
                        "Margin @ Last": margin_last,
                        "Drift vs Saved": drift,
                    })

                if dash_data:
                    dash_df = pd.DataFrame(dash_data)
                    # KPI cards
                    avg_margin_avg = dash_df['Margin @ Avg'].mean()
                    avg_margin_last = dash_df['Margin @ Last'].mean()
                    products_below_30 = (dash_df['Margin @ Avg'] < 30).sum()
                    products_drifting = (dash_df['Drift vs Saved'].abs() > 0.10).sum()
                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Profiles Tracked", f"{len(dash_df)}")
                    k2.metric("Avg Margin (Avg Cost)", f"{avg_margin_avg:.1f}%")
                    k3.metric("Avg Margin (Last Cost)", f"{avg_margin_last:.1f}%", f"{(avg_margin_last - avg_margin_avg):+.1f}pp" if abs(avg_margin_last - avg_margin_avg) > 0.1 else None)
                    k4.metric("⚠️ Need Recalc", f"{products_drifting}", help="Saved COGS differs from current RM-based COGS by more than $0.10")
                    if products_below_30 > 0:
                        st.warning(f"🟠 {products_below_30} product(s) have margin below 30% at average cost. Review pricing.")
                    st.write("---")
                    st.markdown("#### Profile-by-Profile View")
                    st.dataframe(dash_df, use_container_width=True, hide_index=True, column_config={
                        "Retail": st.column_config.NumberColumn(format="$%.2f"),
                        "Saved COGS": st.column_config.NumberColumn(format="$%.2f"),
                        "COGS @ Avg (Now)": st.column_config.NumberColumn(format="$%.2f"),
                        "COGS @ Last (Now)": st.column_config.NumberColumn(format="$%.2f"),
                        "Margin @ Avg": st.column_config.NumberColumn(format="%.1f%%"),
                        "Margin @ Last": st.column_config.NumberColumn(format="%.1f%%"),
                        "Drift vs Saved": st.column_config.NumberColumn(format="$%.2f"),
                    })
                else:
                    st.info("No active COGS profiles yet.")
            else:
                st.info("No COGS profiles saved yet. Go to **Build New COGS** to create one.")

        # ============= TAB 2: BUILD NEW COGS =============
        with tab_build:
            with st.container(border=True):
                st.markdown("#### Step 1: Physical Product Specs")
                c1, c2, c3 = st.columns(3)
                if not formulas_df.empty:
                    f_opts = [f"[{r['fr_code']}] {r['formula_name']}" for _, r in formulas_df.iterrows()]
                    sel_form = c1.selectbox("Base Formula", f_opts)
                else:
                    sel_form = None; c1.warning("No formulas in vault.")
                fill_wt = c2.number_input("Fill Weight per Unit (grams)", min_value=1.0, value=30.0, step=5.0)
                if not packaging.empty:
                    p_opts = [f"[{r['pm_code']}] {r['material_name']}" for _, r in packaging.iterrows()]
                    p_opts.insert(0, "None / Custom")
                    sel_pack = c3.selectbox("Primary Packaging", p_opts)
                else:
                    sel_pack = "None / Custom"; c3.warning("No packaging in vault.")
            with st.container(border=True):
                st.markdown("#### Step 2: Component & Variable Costs (per unit)")
                cm1, cm2, cm3, cm4 = st.columns(4)
                cost_mfg = cm1.number_input("Labor / Mfg ($)", min_value=0.0, value=0.10, step=0.05)
                cost_lbl = cm2.number_input("Label Cost ($)", min_value=0.0, value=0.05, step=0.05)
                cost_sec = cm3.number_input("Secondary Box ($)", min_value=0.0, value=0.00, step=0.05)
                cost_ter = cm4.number_input("Tertiary/Carton ($)", min_value=0.0, value=0.00, step=0.05)

            bulk_cost_avg = 0.0
            bulk_cost_last = 0.0
            n_only = ""
            if sel_form:
                n_only = sel_form.split("] ")[1]
                rec = formulas_df[formulas_df['formula_name'] == n_only].iloc[0]['recipe']
                if isinstance(rec, dict):
                    rec_items = [{"Ingredient": k, "%": v} for k, v in rec.items()]
                elif isinstance(rec, list):
                    rec_items = rec
                else:
                    rec_items = []
                for row in rec_items:
                    ing = row.get('Ingredient'); p = row.get('%', 0)
                    req_g = (p/100) * fill_wt
                    m = inventory[inventory['trade_name'] == ing]
                    if not m.empty:
                        avg_p, last_p = get_rm_costs(m.iloc[0])
                        bulk_cost_avg += (req_g/1000) * avg_p
                        bulk_cost_last += (req_g/1000) * last_p
            bulk_cost = bulk_cost_avg
            pack_cost = 0.0
            if sel_pack != "None / Custom":
                p_only = sel_pack.split("] ")[1]
                pack_cost = float(packaging[packaging['material_name'] == p_only].iloc[0]['cost_per_unit'])
            total_cogs_avg = bulk_cost_avg + pack_cost + cost_mfg + cost_lbl + cost_sec + cost_ter
            total_cogs_last = bulk_cost_last + pack_cost + cost_mfg + cost_lbl + cost_sec + cost_ter
            total_cogs = total_cogs_avg

            # KPI cards
            st.write("---")
            kc1, kc2, kc3 = st.columns(3)
            kc1.metric("COGS @ Avg Cost", f"${total_cogs_avg:.2f}")
            delta_cogs = total_cogs_last - total_cogs_avg
            kc2.metric("COGS @ Last Cost", f"${total_cogs_last:.2f}", f"{'+' if delta_cogs >= 0 else ''}${delta_cogs:.2f}" if abs(delta_cogs) > 0.001 else None)
            kc3.metric("Cost Volatility", f"{((total_cogs_last - total_cogs_avg) / total_cogs_avg * 100) if total_cogs_avg > 0 else 0:.1f}%", help="Difference between Last and Avg cost as % of Avg")

            r1, r2 = st.columns([2, 1])
            with r1:
                st.markdown("#### Cost Breakdown")
                st.dataframe(pd.DataFrame([
                    {"Component": "Formula (Bulk Oil)", "Cost @ Avg": f"${bulk_cost_avg:.4f}", "Cost @ Last": f"${bulk_cost_last:.4f}"},
                    {"Component": "Primary Bottle/Dropper", "Cost @ Avg": f"${pack_cost:.4f}", "Cost @ Last": f"${pack_cost:.4f}"},
                    {"Component": "Labeling", "Cost @ Avg": f"${cost_lbl:.4f}", "Cost @ Last": f"${cost_lbl:.4f}"},
                    {"Component": "Secondary Packaging", "Cost @ Avg": f"${cost_sec:.4f}", "Cost @ Last": f"${cost_sec:.4f}"},
                    {"Component": "Tertiary Packaging", "Cost @ Avg": f"${cost_ter:.4f}", "Cost @ Last": f"${cost_ter:.4f}"},
                    {"Component": "Labor / Mfg Overhead", "Cost @ Avg": f"${cost_mfg:.4f}", "Cost @ Last": f"${cost_mfg:.4f}"}
                ]), use_container_width=True, hide_index=True)
            with r2:
                with st.container(border=True):
                    st.markdown("#### Pricing & Margin")
                    target_retail = st.number_input("Target Retail Price ($)", min_value=0.0, value=total_cogs_avg * 4 if total_cogs_avg > 0 else 0.0, step=1.0, key="build_retail")
                    margin_pct = 0.0
                    margin_pct_last = 0.0
                    if target_retail > 0:
                        gross_profit = target_retail - total_cogs_avg
                        margin_pct = (gross_profit / target_retail) * 100
                        margin_pct_last = ((target_retail - total_cogs_last) / target_retail) * 100
                        m1, m2 = st.columns(2)
                        m1.metric("Margin @ Avg", f"{margin_pct:.1f}%")
                        m2.metric("Margin @ Last", f"{margin_pct_last:.1f}%")

            # What-if simulator
            with st.expander("🎯 What-If Price Simulator"):
                st.write("Drag the slider to explore different retail prices.")
                if total_cogs_avg > 0:
                    sim_min = float(total_cogs_avg * 1.5)
                    sim_max = float(total_cogs_avg * 8)
                    sim_default = float(total_cogs_avg * 4)
                    sim_retail = st.slider("Simulated Retail Price ($)", min_value=sim_min, max_value=sim_max, value=sim_default, step=0.5, key="sim_retail")
                    sim_margin_avg = ((sim_retail - total_cogs_avg) / sim_retail * 100) if sim_retail > 0 else 0
                    sim_margin_last = ((sim_retail - total_cogs_last) / sim_retail * 100) if sim_retail > 0 else 0
                    sim_profit_avg = sim_retail - total_cogs_avg
                    sim_profit_last = sim_retail - total_cogs_last
                    s1, s2, s3, s4 = st.columns(4)
                    s1.metric("Sim Retail", f"${sim_retail:.2f}")
                    s2.metric("Profit @ Avg", f"${sim_profit_avg:.2f}")
                    s3.metric("Margin @ Avg", f"{sim_margin_avg:.1f}%")
                    s4.metric("Margin @ Last", f"{sim_margin_last:.1f}%")
                else:
                    st.info("Build a COGS first (select a formula and packaging) to use the simulator.")

            # Save
            with st.container(border=True):
                st.markdown("#### 💾 Save COGS Profile")
                sc1, sc2 = st.columns([3, 1])
                cogs_name = sc1.text_input("Product Name / SKU", placeholder="e.g., Actiflam 30ml Retail Bottle", key="build_cogs_name")
                sc2.write("<br>", unsafe_allow_html=True)
                if sc2.button("Commit Profile", type="primary", use_container_width=True, key="build_save_btn"):
                    if cogs_name:
                        supabase.table('cogs_records').insert({"product_name": cogs_name, "formula_name": n_only if sel_form else "None", "fill_weight_g": fill_wt, "primary_packaging": sel_pack.split("] ")[1] if sel_pack != "None / Custom" else "Custom", "bulk_cost": bulk_cost, "packaging_cost": pack_cost, "mfg_cost": cost_mfg, "label_cost": cost_lbl, "total_cogs": total_cogs, "target_retail": target_retail, "gross_margin_pct": margin_pct, "version": 1, "is_active": True}).execute()
                        st.success(f"Saved profile: {cogs_name}")
                        clear_cache(); st.rerun()
                    else:
                        st.error("Please enter a Product Name before saving.")

        # ============= TAB 3: SAVED PROFILES =============
        with tab_profiles:
            if not cogs_records_df.empty:
                if 'is_active' in cogs_records_df.columns:
                    active_cogs = cogs_records_df[cogs_records_df['is_active'] != False].copy()
                else:
                    active_cogs = cogs_records_df.copy()
                display_cogs = active_cogs.copy()
                display_cogs['Date'] = pd.to_datetime(display_cogs['created_at'], errors='coerce').dt.strftime('%Y-%m-%d')
                display_cogs.insert(0, '🔍', False)
                with st.container(border=True):
                    edited_cogs = st.data_editor(
                        display_cogs[['🔍', 'Date', 'product_name', 'formula_name', 'fill_weight_g', 'total_cogs', 'target_retail', 'gross_margin_pct']],
                        use_container_width=True, hide_index=True, disabled=['Date', 'formula_name', 'fill_weight_g', 'total_cogs', 'gross_margin_pct'],
                        column_config={"total_cogs": st.column_config.NumberColumn("Total COGS", format="$%.2f"), "target_retail": st.column_config.NumberColumn("Target Retail", format="$%.2f"), "gross_margin_pct": st.column_config.NumberColumn("Margin %", format="%.1f%%")}
                    )
                    if st.button("💾 Synchronize COGS Vault", type="primary"):
                        for index, row in edited_cogs.iterrows():
                            orig = active_cogs.loc[index]
                            if row['product_name'] != orig['product_name'] or row['target_retail'] != orig['target_retail']:
                                new_retail = float(row['target_retail'])
                                new_cogs = float(orig['total_cogs'])
                                new_margin = ((new_retail - new_cogs) / new_retail * 100) if new_retail > 0 else 0.0
                                supabase.table('cogs_records').update({"product_name": row['product_name'], "target_retail": new_retail, "gross_margin_pct": new_margin}).eq('id', int(orig['id'])).execute()
                        st.success("COGS profiles synced!")
                        clear_cache(); st.rerun()
                selected_cogs = edited_cogs[edited_cogs['🔍'] == True]
                if not selected_cogs.empty:
                    cogs_item = active_cogs.loc[selected_cogs.index[0]]
                    st.write("##")
                    with st.container(border=True):
                        st.markdown(f"#### {cogs_item['product_name']}")
                        st.write(f"**Base Formula:** {cogs_item['formula_name']} ({cogs_item['fill_weight_g']}g fill)")
                        st.write(f"**Primary Packaging:** {cogs_item['primary_packaging']}")

                        # --- AT-A-GLANCE COGS COMPARISON ---
                        live_avg, _ = recalc_profile_cogs(cogs_item, "avg")
                        live_last, _ = recalc_profile_cogs(cogs_item, "last")
                        saved_cogs = float(cogs_item['total_cogs'])
                        retail = float(cogs_item['target_retail'])
                        m_saved = ((retail - saved_cogs) / retail * 100) if retail > 0 else 0
                        m_avg = ((retail - live_avg) / retail * 100) if retail > 0 else 0
                        m_last = ((retail - live_last) / retail * 100) if retail > 0 else 0
                        pc1, pc2, pc3, pc4 = st.columns(4)
                        pc1.metric("Retail Price", f"${retail:.2f}")
                        pc2.metric("Saved COGS", f"${saved_cogs:.2f}", f"{m_saved:.1f}% margin")
                        pc3.metric("Live COGS @ Avg", f"${live_avg:.2f}", f"{m_avg:.1f}% margin")
                        pc4.metric("Live COGS @ Last", f"${live_last:.2f}", f"{m_last:.1f}% margin")

                        # --- WHAT-IF SIMULATOR PER PROFILE ---
                        with st.expander("🎯 What-If Price Simulator"):
                            sim_min2 = float(min(live_avg, saved_cogs) * 1.2)
                            sim_max2 = float(max(live_avg, saved_cogs) * 8)
                            sim_retail2 = st.slider("Simulated Retail Price ($)", min_value=sim_min2, max_value=sim_max2, value=retail, step=0.5, key=f"sim_retail_{cogs_item['id']}")
                            sm_avg = ((sim_retail2 - live_avg) / sim_retail2 * 100) if sim_retail2 > 0 else 0
                            sm_last = ((sim_retail2 - live_last) / sim_retail2 * 100) if sim_retail2 > 0 else 0
                            sa1, sa2, sa3 = st.columns(3)
                            sa1.metric("Sim Retail", f"${sim_retail2:.2f}")
                            sa2.metric("Margin @ Avg", f"{sm_avg:.1f}%")
                            sa3.metric("Margin @ Last", f"{sm_last:.1f}%")

                        # --- RECALCULATE COGS ---
                        with st.expander("🔄 Recalculate COGS with Current RM Prices"):
                            recalc_total = live_avg
                            old_total = saved_cogs
                            delta = recalc_total - old_total
                            rc1, rc2, rc3 = st.columns(3)
                            rc1.metric("Saved COGS", f"${old_total:.2f}")
                            rc2.metric("Recalculated (Avg)", f"${recalc_total:.2f}", f"{'↑' if delta > 0 else '↓'} ${abs(delta):.2f}" if abs(delta) > 0.001 else "No change")
                            new_margin = ((retail - recalc_total) / retail * 100) if retail > 0 else 0
                            rc3.metric("New Margin", f"{new_margin:.1f}%")
                            if abs(delta) > 0.001:
                                if st.button("✅ Apply Recalculation (Creates New Version)", type="primary", key="recalc_apply"):
                                    # Recompute bulk separately for accurate save
                                    _, recalc_bulk = recalc_profile_cogs(cogs_item, "avg")
                                    recalc_pack = float(cogs_item.get('packaging_cost', 0) or 0)
                                    recalc_mfg = float(cogs_item.get('mfg_cost', 0) or 0)
                                    recalc_lbl = float(cogs_item.get('label_cost', 0) or 0)
                                    supabase.table('cogs_records').update({"is_active": False}).eq('id', int(cogs_item['id'])).execute()
                                    supabase.table('cogs_records').insert({"product_name": cogs_item['product_name'], "formula_name": cogs_item['formula_name'], "fill_weight_g": float(cogs_item['fill_weight_g']), "primary_packaging": cogs_item['primary_packaging'], "bulk_cost": recalc_bulk, "packaging_cost": recalc_pack, "mfg_cost": recalc_mfg, "label_cost": recalc_lbl, "total_cogs": recalc_total, "target_retail": retail, "gross_margin_pct": new_margin, "version": int(cogs_item.get('version', 1) or 1) + 1, "is_active": True, "parent_id": int(cogs_item['id'])}).execute()
                                    st.success("New COGS version created! Old version archived.")
                                    time.sleep(1); clear_cache(); st.rerun()
                            else:
                                st.info("COGS is already up to date with current RM prices.")

                        # --- VIEW OLDER VERSIONS ---
                        with st.expander("📜 View Older COGS Versions"):
                            all_versions = cogs_records_df[cogs_records_df['product_name'] == cogs_item['product_name']].copy()
                            if 'version' in all_versions.columns:
                                all_versions['version'] = all_versions['version'].fillna(1).astype(int)
                            else:
                                all_versions['version'] = 1
                            if 'is_active' in all_versions.columns:
                                all_versions['is_active'] = all_versions['is_active'].fillna(True)
                            else:
                                all_versions['is_active'] = True
                            all_versions['Status'] = all_versions['is_active'].apply(lambda x: "✅ Active" if x else "📦 Archived")
                            all_versions['Date'] = pd.to_datetime(all_versions['created_at'], errors='coerce').dt.strftime('%Y-%m-%d')
                            if len(all_versions) > 1:
                                st.dataframe(all_versions[['Date', 'version', 'Status', 'total_cogs', 'target_retail', 'gross_margin_pct']].sort_values('version', ascending=False), use_container_width=True, hide_index=True, column_config={"total_cogs": st.column_config.NumberColumn("COGS", format="$%.2f"), "target_retail": st.column_config.NumberColumn("Retail", format="$%.2f"), "gross_margin_pct": st.column_config.NumberColumn("Margin", format="%.1f%%")})
                            else:
                                st.info("No older versions. This is the first version.")

                        with st.expander("System Actions"):
                            del_cogs_pass = st.text_input("Authorization Passcode", type="password", key="dcogsp")
                            if st.button("Erase COGS Profile"):
                                if del_cogs_pass == "lab2026":
                                    supabase.table('cogs_records').delete().eq('id', int(cogs_item['id'])).execute(); clear_cache(); st.rerun()
                                else: st.error("Incorrect passcode.")
            else: st.info("No COGS profiles saved in the vault.")

    # --- STOCK LEVELS (READ-ONLY) ---
    elif menu == "Stock Levels":
        d = load_tables('inventory', 'finished_goods')
        inventory = d['inventory']; finished_goods = d['finished_goods']
        st.title("Stock Levels")
        st.markdown("<p style='opacity: 0.6;'>Read-only view of current inventory.</p>", unsafe_allow_html=True)
        st.markdown("#### Raw Materials")
        if not inventory.empty:
            st.dataframe(inventory[['rm_code', 'trade_name', 'quantity_kg']].rename(columns={'rm_code': 'Code', 'trade_name': 'Material', 'quantity_kg': 'Stock (Kg)'}), use_container_width=True, hide_index=True)
        else:
            st.info("No raw materials registered.")
        st.write("---")
        st.markdown("#### Finished Products")
        if not finished_goods.empty:
            st.dataframe(finished_goods[['fp_code', 'product_name', 'stock_quantity']].rename(columns={'fp_code': 'Code', 'product_name': 'Product', 'stock_quantity': 'In Stock'}), use_container_width=True, hide_index=True)
        else:
            st.info("No finished products in stock.")

    # --- 9. PRODUCTION LOGS ---
    elif menu == "Production Logs":
        st.title("Production Logs")
        st.markdown("<p style='opacity: 0.6;'>GMP-compliant traceability records & Physical Batch Labels.</p>", unsafe_allow_html=True)
        logs = supabase.table('production_records').select("*").order("created_at", desc=True).execute()
        if logs.data:
            df = pd.DataFrame(logs.data)
            df['Date'] = pd.to_datetime(df['created_at'], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')
            disp_logs = df.copy()
            disp_logs.insert(0, '🏷️', False)
            st.write("💡 *Check the box next to any batch to generate its GMP physical labels.*")
            with st.container(border=True):
                edited_logs = st.data_editor(disp_logs[['🏷️', 'id', 'Date', 'batch_number', 'lot_number', 'formula_name', 'batch_size_g', 'total_cost']], use_container_width=True, hide_index=True, disabled=['id', 'Date', 'batch_number', 'lot_number', 'formula_name', 'batch_size_g', 'total_cost'], column_config={"id": None})
            sel_logs = edited_logs[edited_logs['🏷️'] == True]
            if not sel_logs.empty:
                s_log = df[df['id'] == sel_logs.iloc[0]['id']].iloc[0]
                st.write("##")
                with st.container(border=True):
                    st.markdown(f"#### 🖨️ Label Generator: {s_log['batch_number']}")
                    st.write(f"**Formula:** {s_log['formula_name']} | **Lot:** {s_log['lot_number']} | **Size:** {s_log['batch_size_g']}g")
                    pdf_bytes = generate_batch_labels_pdf(s_log['formula_name'], s_log['batch_number'], s_log['lot_number'], pd.to_datetime(s_log['created_at'], errors='coerce').strftime('%Y-%m-%d'))
                    st.download_button(label="📄 Download GMP Batch Label Sheet (PDF)", data=pdf_bytes, file_name=f"Labels_{s_log['batch_number']}.pdf", mime="application/pdf", use_container_width=True, type="primary")
        else: 
            st.info("No records found in the vault.")

    # --- DATA CLEANING ---
    elif menu == "Data Cleaning":
        d = load_tables('sales_records', 'finished_goods', 'consignment', 'cogs_records', 'clients')
        sales_records_df = d['sales_records']; finished_goods = d['finished_goods']; consignment_df = d['consignment']; cogs_records_df = d['cogs_records']; clients_df = d['clients']
        st.title("Data Cleaning & Consolidation")
        st.markdown("<p style='opacity: 0.6;'>Merge duplicate product names or client names across all tables. Changes are permanent and affect Sales, COGS, Finished Products, Consignment, and Client records.</p>", unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["🏷️ Product Name Consolidation", "👤 Client Name Consolidation"])

        with tab1:
            st.markdown("#### Consolidate Product Names")
            st.info("💡 This will rename a product across ALL tables: Sales Records, Finished Products, Consignment, and COGS Profiles.")
            all_product_names = set()
            if not sales_records_df.empty:
                all_product_names.update(sales_records_df['order_description'].dropna().unique())
            if not finished_goods.empty:
                all_product_names.update(finished_goods['product_name'].dropna().unique())
            if not consignment_df.empty:
                all_product_names.update(consignment_df['product_name'].dropna().unique())
            if not cogs_records_df.empty:
                all_product_names.update(cogs_records_df['product_name'].dropna().unique())
            all_product_names = sorted(all_product_names)
            if all_product_names:
                st.markdown("**All product names found across your system:**")
                st.dataframe(pd.DataFrame({"Product Name": all_product_names, "#": range(1, len(all_product_names) + 1)})[['#', 'Product Name']], use_container_width=True, hide_index=True)
                st.write("---")
                with st.form("merge_products"):
                    st.markdown("**Merge product names**")
                    mp1, mp2 = st.columns(2)
                    old_name = mp1.selectbox("Rename this (old name):", all_product_names, key="old_prod")
                    new_name = mp2.selectbox("To this (correct name):", all_product_names, key="new_prod")
                    custom_name = st.text_input("Or type a completely new name (overrides selection above):", key="custom_prod")
                    if st.form_submit_button("Apply Rename Across All Tables", type="primary"):
                        final_name = custom_name.strip() if custom_name.strip() else new_name
                        if old_name == final_name:
                            st.error("Old and new names are the same.")
                        else:
                            changes = 0
                            if not sales_records_df.empty:
                                matches = sales_records_df[sales_records_df['order_description'] == old_name]
                                for _, row in matches.iterrows():
                                    supabase.table('sales_records').update({"order_description": final_name}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            if not finished_goods.empty:
                                matches = finished_goods[finished_goods['product_name'] == old_name]
                                for _, row in matches.iterrows():
                                    supabase.table('finished_products').update({"product_name": final_name}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            if not consignment_df.empty:
                                matches = consignment_df[consignment_df['product_name'] == old_name]
                                for _, row in matches.iterrows():
                                    supabase.table('consignment_records').update({"product_name": final_name}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            if not cogs_records_df.empty:
                                matches = cogs_records_df[cogs_records_df['product_name'] == old_name]
                                for _, row in matches.iterrows():
                                    supabase.table('cogs_records').update({"product_name": final_name}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            pf_df = fetch_vault_data('portfolios', 'portfolio_name')
                            if not pf_df.empty:
                                for _, pf_row in pf_df.iterrows():
                                    pf_prods = pf_row['products'] if isinstance(pf_row['products'], list) else []
                                    if old_name in pf_prods:
                                        updated_prods = list(dict.fromkeys([final_name if p == old_name else p for p in pf_prods]))
                                        supabase.table('portfolios').update({"products": updated_prods}).eq('id', int(pf_row['id'])).execute()
                                        changes += 1
                            st.success(f"Renamed '{old_name}' → '{final_name}' across {changes} records.")
                            time.sleep(1.5); clear_cache(); st.rerun()
            else:
                st.info("No product names found in the system.")

        with tab2:
            st.markdown("#### Consolidate Client / Account Names")
            st.info("💡 This will rename a client across Sales Records, Client Database, and Consignment partner names.")
            all_client_names = set()
            if not sales_records_df.empty:
                all_client_names.update(sales_records_df['account'].dropna().unique())
            if not clients_df.empty:
                all_client_names.update(clients_df['client_name'].dropna().unique())
            if not consignment_df.empty:
                all_client_names.update(consignment_df['partner_name'].dropna().unique())
            all_client_names = sorted(all_client_names)
            if all_client_names:
                st.markdown("**All client/partner names found across your system:**")
                st.dataframe(pd.DataFrame({"Client Name": all_client_names, "#": range(1, len(all_client_names) + 1)})[['#', 'Client Name']], use_container_width=True, hide_index=True)
                st.write("---")
                with st.form("merge_clients"):
                    st.markdown("**Merge client names**")
                    mc1, mc2 = st.columns(2)
                    old_client = mc1.selectbox("Rename this (old name):", all_client_names, key="old_client")
                    new_client = mc2.selectbox("To this (correct name):", all_client_names, key="new_client")
                    custom_client = st.text_input("Or type a completely new name (overrides selection above):", key="custom_client")
                    if st.form_submit_button("Apply Rename Across All Tables", type="primary"):
                        final_client = custom_client.strip() if custom_client.strip() else new_client
                        if old_client == final_client:
                            st.error("Old and new names are the same.")
                        else:
                            changes = 0
                            if not sales_records_df.empty:
                                matches = sales_records_df[sales_records_df['account'] == old_client]
                                for _, row in matches.iterrows():
                                    supabase.table('sales_records').update({"account": final_client}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            if not clients_df.empty:
                                matches = clients_df[clients_df['client_name'] == old_client]
                                for _, row in matches.iterrows():
                                    supabase.table('clients').update({"client_name": final_client}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            if not consignment_df.empty:
                                matches = consignment_df[consignment_df['partner_name'] == old_client]
                                for _, row in matches.iterrows():
                                    supabase.table('consignment_records').update({"partner_name": final_client}).eq('id', int(row['id'])).execute()
                                    changes += 1
                            # Delete duplicate client records with old name
                            if not clients_df.empty:
                                dupes = clients_df[clients_df['client_name'] == old_client]
                                existing_new = clients_df[clients_df['client_name'] == final_client]
                                if not existing_new.empty and not dupes.empty:
                                    for _, drow in dupes.iterrows():
                                        supabase.table('clients').delete().eq('id', int(drow['id'])).execute()
                                        changes += 1
                            st.success(f"Renamed '{old_client}' → '{final_client}' across {changes} records.")
                            time.sleep(1.5); clear_cache(); st.rerun()
            else:
                st.info("No client names found in the system.")

    # --- PORTFOLIO BUILDER ---
    elif menu == "Portfolio Builder":
        d = load_tables('sales_records', 'finished_goods', 'cogs_records', 'consignment')
        sales_records_df = d['sales_records']; finished_goods = d['finished_goods']; cogs_records_df = d['cogs_records']; consignment_df = d['consignment']
        st.title("Portfolio Builder")
        st.markdown("<p style='opacity: 0.6;'>Group related products into portfolios to track their combined performance in Analytics.</p>", unsafe_allow_html=True)

        # Load portfolios from session or supabase
        portfolios_df = fetch_vault_data('portfolios', 'portfolio_name')

        # --- Create New Portfolio ---
        with st.expander("➕ Create New Portfolio"):
            all_product_names = set()
            if not sales_records_df.empty:
                all_product_names.update(sales_records_df['order_description'].dropna().unique())
            if not finished_goods.empty:
                all_product_names.update(finished_goods['product_name'].dropna().unique())
            if not cogs_records_df.empty:
                all_product_names.update(cogs_records_df['product_name'].dropna().unique())
            all_product_names = sorted(all_product_names)
            if all_product_names:
                with st.form("create_portfolio"):
                    pf_name = st.text_input("Portfolio Name", placeholder="e.g., Massage Candle Line, Actiflam Family")
                    pf_products = st.multiselect("Select products in this portfolio:", all_product_names)
                    pf_desc = st.text_input("Description (optional)", placeholder="e.g., All candle SKUs across fragrances")
                    if st.form_submit_button("Create Portfolio", type="primary"):
                        if not pf_name.strip():
                            st.error("Please enter a portfolio name.")
                        elif len(pf_products) < 2:
                            st.error("Select at least 2 products to group.")
                        else:
                            supabase.table('portfolios').insert({"portfolio_name": pf_name.strip(), "products": list(dict.fromkeys(pf_products)), "description": pf_desc}).execute()
                            st.success(f"Portfolio '{pf_name}' created with {len(pf_products)} products!")
                            time.sleep(1); clear_cache(); st.rerun()

        # --- View Existing Portfolios ---
        st.write("---")
        st.markdown("#### Saved Portfolios")
        if not portfolios_df.empty:
            for idx, pf in portfolios_df.iterrows():
                products_list = pf['products'] if isinstance(pf['products'], list) else []
                with st.container(border=True):
                    pc1, pc2 = st.columns([3, 1])
                    pc1.markdown(f"**{pf['portfolio_name']}**")
                    if pf.get('description'):
                        pc1.write(f"*{pf['description']}*")
                    pc1.write(f"Products: {', '.join(products_list)}")
                    # Performance preview
                    if not sales_records_df.empty and products_list:
                        pf_sales = sales_records_df[sales_records_df['order_description'].isin(products_list)]
                        if not pf_sales.empty:
                            pf_rev = pf_sales['gross_revenue'].sum()
                            pf_units = pf_sales['quantity'].sum()
                            pf_profit = pf_sales['net_profit'].sum()
                            pc2.metric("Revenue", f"${pf_rev:,.2f}")
                            pc1.write(f"📊 {int(pf_units)} units sold | ${pf_profit:,.2f} profit")
                    # Edit / Delete
                    with st.expander("Manage"):
                        valid_defaults = [p for p in products_list if p in all_product_names]
                        if len(valid_defaults) < len(products_list):
                            st.warning(f"⚠️ {len(products_list) - len(valid_defaults)} product(s) in this portfolio no longer exist. They were likely renamed in Data Cleaning.")
                        edit_products = st.multiselect(f"Edit products in '{pf['portfolio_name']}':", all_product_names, default=valid_defaults, key=f"edit_pf_{pf['id']}")
                        ec1, ec2 = st.columns(2)
                        if ec1.button("Update Products", key=f"upd_pf_{pf['id']}"):
                            supabase.table('portfolios').update({"products": list(dict.fromkeys(edit_products))}).eq('id', int(pf['id'])).execute()
                            st.success("Portfolio updated!")
                            clear_cache(); st.rerun()
                        if ec2.button("Delete Portfolio", key=f"del_pf_{pf['id']}"):
                            supabase.table('portfolios').delete().eq('id', int(pf['id'])).execute()
                            clear_cache(); st.rerun()
        else:
            st.info("No portfolios created yet. Use the form above to group related products.")

    # --- PRICE MANAGER ---
    elif menu == "Price Manager":
        d = load_tables('finished_goods', 'cogs_records', 'sales_records', 'consignment')
        finished_goods = d['finished_goods']; cogs_records_df = d['cogs_records']; sales_records_df = d['sales_records']; consignment_df = d['consignment']
        st.title("Price Manager")
        st.markdown("<p style='opacity: 0.6;'>Overview of all pricing across the system. Flags discrepancies between COGS profiles, Finished Products, and actual sales prices.</p>", unsafe_allow_html=True)

        # --- 1. Master Price Table ---
        st.markdown("#### Master Price Overview")
        price_data = []
        # Collect ALL unique product names across FP, Sales, COGS, Consignment
        all_products = set()
        if not finished_goods.empty:
            all_products.update(finished_goods['product_name'].dropna().unique())
        if not sales_records_df.empty:
            all_products.update(sales_records_df['order_description'].dropna().unique())
        if not cogs_records_df.empty:
            all_products.update(cogs_records_df['product_name'].dropna().unique())
        all_products = sorted(all_products)
        for prod_name in all_products:
            row = {"Product": prod_name, "FP Retail": None, "FP COGS": None, "FP Margin %": None, "COGS Profile Price": None, "COGS Profile COGS": None, "Avg Sold Price": None, "Min Sold Price": None, "Max Sold Price": None, "Flags": []}
            # Match to Finished Products
            fp_match = finished_goods[finished_goods['product_name'] == prod_name] if not finished_goods.empty else pd.DataFrame()
            if not fp_match.empty:
                fp = fp_match.iloc[0]
                row["FP Retail"] = float(fp['retail_price'])
                row["FP COGS"] = float(fp['unit_cogs'])
                if fp['retail_price'] > 0:
                    row["FP Margin %"] = round(((fp['retail_price'] - fp['unit_cogs']) / fp['retail_price']) * 100, 1)
            else:
                row["Flags"].append("ℹ️ Not in Finished Products")
            # Match to COGS profile
            if not cogs_records_df.empty:
                if 'is_active' in cogs_records_df.columns:
                    active_cogs = cogs_records_df[cogs_records_df['is_active'] != False]
                else:
                    active_cogs = cogs_records_df
                cogs_match = active_cogs[active_cogs['product_name'] == prod_name]
                if not cogs_match.empty:
                    cogs_row = cogs_match.iloc[0]
                    row["COGS Profile Price"] = float(cogs_row['target_retail'])
                    row["COGS Profile COGS"] = float(cogs_row['total_cogs'])
                    if row["FP Retail"] is not None and abs(float(cogs_row['target_retail']) - row["FP Retail"]) > 0.01:
                        row["Flags"].append("⚠️ FP retail ≠ COGS target retail")
                    if row["FP COGS"] is not None and abs(float(cogs_row['total_cogs']) - row["FP COGS"]) > 0.01:
                        row["Flags"].append("⚠️ FP unit COGS ≠ COGS profile COGS")
                else:
                    row["Flags"].append("❌ No COGS profile")
            # Match to actual sales
            if not sales_records_df.empty:
                sold = sales_records_df[sales_records_df['order_description'] == prod_name]
                if not sold.empty:
                    row["Avg Sold Price"] = round(float(sold['unit_price'].mean()), 2)
                    row["Min Sold Price"] = round(float(sold['unit_price'].min()), 2)
                    row["Max Sold Price"] = round(float(sold['unit_price'].max()), 2)
                    ref_cogs = row["FP COGS"] or row["COGS Profile COGS"]
                    ref_retail = row["FP Retail"] or row["COGS Profile Price"]
                    if ref_cogs and row["Avg Sold Price"] < ref_cogs:
                        row["Flags"].append("🔴 Avg sold price BELOW COGS")
                    if ref_cogs and row["Min Sold Price"] < ref_cogs:
                        row["Flags"].append("🟠 Some sales below COGS")
                    if ref_retail and row["Max Sold Price"] > ref_retail * 1.5:
                        row["Flags"].append("🟡 Some sales >150% of retail")
                else:
                    row["Flags"].append("ℹ️ Never sold")
            # Low margin flag
            if row["FP Margin %"] is not None and row["FP Margin %"] < 30:
                row["Flags"].append("🟠 Margin below 30%")
            row["Flags"] = " | ".join(row["Flags"]) if row["Flags"] else "✅ OK"
            price_data.append(row)
        if price_data:
            price_df = pd.DataFrame(price_data)
            # Show flagged items first
            price_df['has_issue'] = price_df['Flags'].apply(lambda x: 0 if x == "✅ OK" else 1)
            price_df = price_df.sort_values(['has_issue', 'Product'], ascending=[False, True])
            st.dataframe(price_df[['Product', 'FP Retail', 'FP COGS', 'FP Margin %', 'COGS Profile Price', 'COGS Profile COGS', 'Avg Sold Price', 'Min Sold Price', 'Max Sold Price', 'Flags']], use_container_width=True, hide_index=True, column_config={
                "FP Retail": st.column_config.NumberColumn(format="$%.2f"),
                "FP COGS": st.column_config.NumberColumn(format="$%.2f"),
                "FP Margin %": st.column_config.NumberColumn(format="%.1f%%"),
                "COGS Profile Price": st.column_config.NumberColumn(format="$%.2f"),
                "COGS Profile COGS": st.column_config.NumberColumn(format="$%.2f"),
                "Avg Sold Price": st.column_config.NumberColumn(format="$%.2f"),
                "Min Sold Price": st.column_config.NumberColumn(format="$%.2f"),
                "Max Sold Price": st.column_config.NumberColumn(format="$%.2f"),
            })
            # Count issues
            issues = [r for r in price_data if r['Flags'] != "✅ OK"]
            if issues:
                st.warning(f"⚠️ {len(issues)} product(s) have pricing discrepancies. Review the flags above.")
            else:
                st.success("✅ All pricing is consistent across the system.")

            # --- 2. Bulk Price Editor ---
            st.write("---")
            st.markdown("#### Bulk Retail Price Update")
            st.info("💡 Edit retail prices below. Changes will update both Finished Products and the active COGS profile.")
            edit_price_data = []
            for _, fp in finished_goods.iterrows():
                edit_price_data.append({"id": int(fp['id']), "Product": fp['product_name'], "Current Retail": float(fp['retail_price']), "New Retail": float(fp['retail_price']), "Unit COGS": float(fp['unit_cogs'])})
            edit_price_df = pd.DataFrame(edit_price_data)
            edited_prices = st.data_editor(edit_price_df, use_container_width=True, hide_index=True, disabled=['id', 'Product', 'Current Retail', 'Unit COGS'], column_config={
                "id": None,
                "Current Retail": st.column_config.NumberColumn(format="$%.2f"),
                "New Retail": st.column_config.NumberColumn(format="$%.2f", min_value=0.0),
                "Unit COGS": st.column_config.NumberColumn(format="$%.2f"),
            })
            if st.button("💾 Apply Price Changes", type="primary"):
                changes = 0
                for idx, row in edited_prices.iterrows():
                    if abs(row['New Retail'] - row['Current Retail']) > 0.001:
                        new_retail = float(row['New Retail'])
                        unit_cogs = float(row['Unit COGS'])
                        new_margin = ((new_retail - unit_cogs) / new_retail * 100) if new_retail > 0 else 0.0
                        # Update Finished Products
                        supabase.table('finished_products').update({"retail_price": new_retail}).eq('id', int(row['id'])).execute()
                        # Update active COGS profile
                        if not cogs_records_df.empty:
                            if 'is_active' in cogs_records_df.columns:
                                active_cogs = cogs_records_df[cogs_records_df['is_active'] != False]
                            else:
                                active_cogs = cogs_records_df
                            cogs_match = active_cogs[active_cogs['product_name'] == row['Product']]
                            if not cogs_match.empty:
                                supabase.table('cogs_records').update({"target_retail": new_retail, "gross_margin_pct": new_margin}).eq('id', int(cogs_match.iloc[0]['id'])).execute()
                        changes += 1
                if changes > 0:
                    st.success(f"Updated {changes} product price(s)!")
                    clear_cache(); st.rerun()
                else:
                    st.info("No prices were changed.")

            # --- 3. Consignment Price Check ---
            if not consignment_df.empty:
                st.write("---")
                st.markdown("#### Consignment Price Check")
                active_cons = consignment_df[consignment_df['status'] == 'Active'].copy() if 'status' in consignment_df.columns else consignment_df.copy()
                if not active_cons.empty:
                    cons_price_data = []
                    for _, c in active_cons.iterrows():
                        fp_match = finished_goods[finished_goods['product_name'] == c['product_name']]
                        current_retail = float(fp_match.iloc[0]['retail_price']) if not fp_match.empty else None
                        flags = []
                        if current_retail and float(c['wholesale_price']) >= current_retail:
                            flags.append("🔴 Partner price ≥ retail price")
                        if current_retail and float(c['wholesale_price']) < float(c.get('unit_cogs', 0) or 0):
                            flags.append("🔴 Partner price below COGS")
                        cons_price_data.append({"Partner": c['partner_name'], "Product": c['product_name'], "Ref": c['order_ref_number'], "Partner Price": float(c['wholesale_price']), "Retail Price": float(c['retail_price']), "Current FP Retail": current_retail, "Flags": " | ".join(flags) if flags else "✅ OK"})
                    cons_price_df = pd.DataFrame(cons_price_data)
                    st.dataframe(cons_price_df, use_container_width=True, hide_index=True, column_config={
                        "Partner Price": st.column_config.NumberColumn(format="$%.2f"),
                        "Retail Price": st.column_config.NumberColumn(format="$%.2f"),
                        "Current FP Retail": st.column_config.NumberColumn(format="$%.2f"),
                    })
        else:
            st.info("No finished products in the system to manage pricing for.")

    # --- BULK IMPORT ---
    elif menu == "Bulk Import":
        d = load_tables('inventory', 'packaging')
        inventory = d['inventory']; packaging = d['packaging']
        st.title("Bulk Import Materials")
        st.markdown("<p style='opacity: 0.6;'>Add multiple raw materials and packaging items at once. Type rows directly or upload a CSV.</p>", unsafe_allow_html=True)

        st.info("💡 You only need 3 fields: **Name, Qty, Price**. The system auto-detects whether it's RM (Kg) or Packaging (units), uses today's date, and auto-generates lot numbers. New items get auto-created codes.")

        # Helpers
        def find_match(candidate, existing_names, threshold=0.8):
            if not existing_names:
                return None, 0.0
            from difflib import SequenceMatcher
            best_name = None
            best_ratio = 0.0
            cand_lower = str(candidate).strip().lower()
            for name in existing_names:
                if pd.isna(name): continue
                ratio = SequenceMatcher(None, cand_lower, str(name).strip().lower()).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_name = name
            return (best_name, best_ratio) if best_ratio >= threshold else (None, best_ratio)

        rm_names = inventory['trade_name'].tolist() if not inventory.empty else []
        pm_names = packaging['material_name'].tolist() if not packaging.empty else []

        tab_table, tab_csv = st.tabs(["✏️ Type/Paste Rows", "📁 Upload File"])

        # --- TAB 1: In-app table ---
        import_df = None
        with tab_table:
            st.markdown("**Type or paste rows below.** You can also copy/paste from Excel.")
            blank_rows = pd.DataFrame([{"name": "", "qty": 0.0, "price": 0.0} for _ in range(5)])
            edited_rows = st.data_editor(blank_rows, num_rows="dynamic", use_container_width=True, hide_index=True, key="bulk_input_table", column_config={
                "name": st.column_config.TextColumn("Name", width="large"),
                "qty": st.column_config.NumberColumn("Quantity", min_value=0.0, step=0.5),
                "price": st.column_config.NumberColumn("Price ($)", format="$%.2f", min_value=0.0)
            })
            valid_rows = edited_rows[(edited_rows['name'].astype(str).str.strip() != "") & (edited_rows['qty'] > 0)]
            if not valid_rows.empty and st.button("📋 Preview & Confirm", key="preview_table_btn", type="primary"):
                import_df = valid_rows.copy()
                st.session_state.bulk_preview_df = import_df

        # --- TAB 2: Excel/CSV upload ---
        with tab_csv:
            st.markdown("**Download a smart Excel template** with dropdowns of all your existing materials and auto-filled prices.")

            # Generate Excel template with data validation
            def build_excel_template():
                from openpyxl import Workbook
                from openpyxl.worksheet.datavalidation import DataValidation
                from openpyxl.styles import Font, PatternFill, Alignment
                import io
                wb = Workbook()
                ws = wb.active
                ws.title = "Bulk Import"
                # Header
                ws['A1'] = "name"; ws['B1'] = "qty"; ws['C1'] = "price"
                for cell in ['A1', 'B1', 'C1']:
                    ws[cell].font = Font(bold=True, color="FFFFFF")
                    ws[cell].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                    ws[cell].alignment = Alignment(horizontal="center")
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 14
                # Reference sheet with name -> last price lookup
                ref_sheet = wb.create_sheet("_Reference")
                ref_sheet['A1'] = "Name"; ref_sheet['B1'] = "Last Price"
                ref_row = 2
                all_refs = []
                for _, mat in inventory.iterrows():
                    ref_sheet.cell(row=ref_row, column=1, value=str(mat['trade_name']))
                    ref_sheet.cell(row=ref_row, column=2, value=float(mat['price_per_kg']))
                    all_refs.append(str(mat['trade_name']))
                    ref_row += 1
                for _, pm in packaging.iterrows():
                    ref_sheet.cell(row=ref_row, column=1, value=str(pm['material_name']))
                    ref_sheet.cell(row=ref_row, column=2, value=float(pm['cost_per_unit']))
                    all_refs.append(str(pm['material_name']))
                    ref_row += 1
                ref_sheet.sheet_state = 'hidden'
                # Data validation: dropdown for column A
                if all_refs:
                    last_ref_row = ref_row - 1
                    dv = DataValidation(type="list", formula1=f"=_Reference!$A$2:$A${last_ref_row}", allow_blank=True)
                    dv.add("A2:A1000")
                    ws.add_data_validation(dv)
                    # Auto-fill price using VLOOKUP for column C
                    for r in range(2, 1001):
                        ws.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(A{r},_Reference!A:B,2,FALSE),"")')
                # Empty row guidance
                ws['A2'] = ""; ws['B2'] = ""
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            try:
                excel_bytes = build_excel_template()
                st.download_button("📥 Download Smart Excel Template (with dropdowns)", data=excel_bytes, file_name="bulk_import_template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_xlsx", type="primary")
            except Exception as e:
                st.warning(f"Excel template error: {e}. CSV fallback still available below.")

            sample_csv = "name,qty,price\nRosemary Oil,2.5,45.00\nMenthol Crystal,1.0,80.00\n100mL Amber Bottle,200,0.35\n"
            st.download_button("📥 Download Plain CSV Template (fallback)", data=sample_csv, file_name="bulk_import_template.csv", mime="text/csv", key="dl_template")

            st.markdown("---")
            st.markdown("**Upload your filled file (Excel or CSV)**")
            uploaded = st.file_uploader("Upload", type=['csv', 'xlsx'], key="bulk_upload")
            if uploaded is not None:
                try:
                    if uploaded.name.endswith('.xlsx'):
                        csv_df = pd.read_excel(uploaded, sheet_name=0)
                    else:
                        csv_df = pd.read_csv(uploaded)
                    required = ['name', 'qty', 'price']
                    missing = [c for c in required if c not in csv_df.columns]
                    if missing:
                        st.error(f"❌ Missing columns: {', '.join(missing)}")
                    else:
                        # Drop empty rows
                        csv_df = csv_df.dropna(subset=['name', 'qty'])
                        csv_df = csv_df[csv_df['name'].astype(str).str.strip() != ""]
                        if csv_df.empty:
                            st.warning("No valid rows found in the file.")
                        else:
                            st.session_state.bulk_preview_df = csv_df[required].reset_index(drop=True)
                            st.success(f"Loaded {len(csv_df)} rows. Scroll down to preview.")
                except Exception as e:
                    st.error(f"Error: {e}")

        # --- Preview & Confirm ---
        if "bulk_preview_df" in st.session_state and not st.session_state.bulk_preview_df.empty:
            st.write("---")
            st.markdown("#### 🔍 Preview & Confirm")
            preview_data = st.session_state.bulk_preview_df

            preview_rows = []
            for idx, row in preview_data.iterrows():
                name = str(row['name']).strip()
                qty = float(row['qty'])
                price = float(row['price'])
                # Try matching against both RM and PM
                rm_match, rm_score = find_match(name, rm_names)
                pm_match, pm_score = find_match(name, pm_names)
                # Determine type and target
                if rm_score > pm_score and rm_match:
                    detected_type = "RM"
                    if rm_match.strip().lower() == name.strip().lower():
                        action = "ADD LOT (exact)"
                    else:
                        action = f"ADD LOT (fuzzy {int(rm_score*100)}%)"
                    target = rm_match
                elif pm_match:
                    detected_type = "PM"
                    if pm_match.strip().lower() == name.strip().lower():
                        action = "ADD LOT (exact)"
                    else:
                        action = f"ADD LOT (fuzzy {int(pm_score*100)}%)"
                    target = pm_match
                else:
                    # No match - guess based on context (RM if has decimals, PM if whole number ≥ 10)
                    detected_type = "RM" if qty < 10 or qty != int(qty) else "PM"
                    action = "CREATE NEW"
                    target = name
                preview_rows.append({
                    "Row": idx + 1, "Type": detected_type, "Input Name": name, "Action": action,
                    "Will Match To": target, "Qty": qty, "Price": price
                })
            preview_df = pd.DataFrame(preview_rows)
            preview_df.insert(0, 'Confirm', True)

            # Allow user to override Type for ambiguous rows
            edited_preview = st.data_editor(preview_df, use_container_width=True, hide_index=True, disabled=['Row', 'Input Name', 'Action', 'Will Match To', 'Qty', 'Price'], column_config={
                "Confirm": st.column_config.CheckboxColumn("Import?", default=True),
                "Type": st.column_config.SelectboxColumn("Type", options=["RM", "PM"], required=True)
            })

            st.warning("⚠️ Review carefully. Adjust **Type** column if the auto-detection is wrong. Uncheck rows you don't want.")
            cb1, cb2 = st.columns(2)
            if cb1.button("🚀 Confirm & Import", type="primary"):
                confirmed = edited_preview[edited_preview['Confirm'] == True]
                imports_done = 0
                new_created = 0
                today_str = datetime.today().strftime('%Y-%m-%d')
                for _, prow in confirmed.iterrows():
                    csv_row = preview_data.iloc[prow['Row'] - 1]
                    row_type = prow['Type']
                    target_name = prow['Will Match To']
                    qty = float(csv_row['qty'])
                    price = float(csv_row['price'])
                    is_new = prow['Action'] == "CREATE NEW"

                    if row_type == "RM":
                        if is_new:
                            next_id = 1 if inventory.empty else int(inventory['id'].max()) + 1
                            rm_code = f"RM{next_id:05d}"
                            lot_num = f"{rm_code}-L01"
                            exp_str = (datetime.today() + pd.DateOffset(years=2)).strftime('%Y-%m-%d')
                            init_lot = [{"Lot Number": lot_num, "Mfg Date": today_str, "Rcv Date": today_str, "Exp Date": exp_str, "Qty (Kg)": qty, "Price/Kg": price, "Current": True}]
                            supabase.table('inventory').insert({"rm_code": rm_code, "trade_name": target_name, "inci_name": "", "price_per_kg": price, "quantity_kg": qty, "lots": init_lot}).execute()
                            new_created += 1
                            inventory = pd.concat([inventory, pd.DataFrame([{"id": next_id, "rm_code": rm_code, "trade_name": target_name, "quantity_kg": qty, "price_per_kg": price}])], ignore_index=True) if not inventory.empty else pd.DataFrame([{"id": next_id, "rm_code": rm_code, "trade_name": target_name, "quantity_kg": qty, "price_per_kg": price}])
                        else:
                            mat = inventory[inventory['trade_name'] == target_name].iloc[0]
                            lots = mat.get('lots', [])
                            if isinstance(lots, float) or (isinstance(lots, str) and lots in ["", "nan", "[]"]): lots = []
                            for l in lots: l['Current'] = False
                            lot_count = len(lots) + 1
                            lot_num = f"{mat['rm_code']}-L{lot_count:02d}"
                            exp_str = (datetime.today() + pd.DateOffset(years=2)).strftime('%Y-%m-%d')
                            lots.append({"Lot Number": lot_num, "Mfg Date": today_str, "Rcv Date": today_str, "Exp Date": exp_str, "Qty (Kg)": qty, "Price/Kg": price, "Current": True})
                            new_total = float(mat['quantity_kg']) + qty
                            supabase.table('inventory').update({"lots": lots, "quantity_kg": new_total, "price_per_kg": price}).eq('id', int(mat['id'])).execute()
                        imports_done += 1
                    elif row_type == "PM":
                        if is_new:
                            next_id = 1 if packaging.empty else int(packaging['id'].max()) + 1
                            pm_code = f"PM{next_id:05d}"
                            lot_num = f"{pm_code}-L01"
                            init_lot = [{"Lot Number": lot_num, "Rcv Date": today_str, "Qty (Units)": int(qty), "Current": True}]
                            supabase.table('packaging').insert({"pm_code": pm_code, "material_name": target_name, "supplier": "", "cost_per_unit": price, "remaining_quantity": int(qty), "lots": init_lot}).execute()
                            new_created += 1
                            packaging = pd.concat([packaging, pd.DataFrame([{"id": next_id, "pm_code": pm_code, "material_name": target_name, "remaining_quantity": int(qty), "cost_per_unit": price}])], ignore_index=True) if not packaging.empty else pd.DataFrame([{"id": next_id, "pm_code": pm_code, "material_name": target_name, "remaining_quantity": int(qty), "cost_per_unit": price}])
                        else:
                            pm = packaging[packaging['material_name'] == target_name].iloc[0]
                            lots = pm.get('lots', [])
                            if isinstance(lots, float) or (isinstance(lots, str) and lots in ["", "nan", "[]"]): lots = []
                            for l in lots: l['Current'] = False
                            lot_count = len(lots) + 1
                            lot_num = f"{pm['pm_code']}-L{lot_count:02d}"
                            lots.append({"Lot Number": lot_num, "Rcv Date": today_str, "Qty (Units)": int(qty), "Current": True})
                            new_total = int(pm['remaining_quantity']) + int(qty)
                            supabase.table('packaging').update({"lots": lots, "remaining_quantity": new_total, "cost_per_unit": price}).eq('id', int(pm['id'])).execute()
                        imports_done += 1
                st.success(f"✅ Imported {imports_done} row(s). Created {new_created} new material(s).")
                if "cogs_synced_this_session" in st.session_state:
                    del st.session_state.cogs_synced_this_session
                del st.session_state.bulk_preview_df
                time.sleep(2); clear_cache(); st.rerun()
            if cb2.button("❌ Cancel / Reset", key="cancel_bulk"):
                del st.session_state.bulk_preview_df
                st.rerun()
